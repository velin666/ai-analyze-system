#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test the complete Excel modification flow with res.md and res2.md
"""
import sys
import os
sys.path.insert(0, 'server/api/files')
from modify_excel import modify_excel

# Read test markdown files
with open('server/api/res.md', 'r', encoding='utf-8') as f:
    res_content = f.read()

with open('server/api/res2.md', 'r', encoding='utf-8') as f:
    res2_content = f.read()

# Create a test Excel file
from openpyxl import Workbook

def create_test_excel(filename):
    """Create a test Excel file with proper structure"""
    wb = Workbook()
    ws = wb.active
    
    # Add some header rows
    ws['A1'] = 'XRD'
    ws['B1'] = 'HFC51-SD01 燃气烘烤炉      电气元件清单'
    ws['C1'] = '文件编号：JS-03'
    
    ws['A2'] = ''
    ws['A3'] = '项目说明：测试项目'
    
    # Row 4 is the actual header
    ws['A4'] = '序号'
    ws['B4'] = 'ERP识别码'
    ws['C4'] = '品牌'
    ws['D4'] = '名称'
    ws['E4'] = '型号尺寸'
    ws['F4'] = '数量'
    ws['G4'] = '单位'
    ws['H4'] = '备注'
    
    # Add some test data
    ws['A5'] = '4'
    ws['B5'] = 'OLD-004'
    ws['C5'] = '西门子'
    ws['D5'] = '触摸屏'
    ws['E5'] = 'OLD-MODEL'
    ws['F5'] = '1'
    ws['G5'] = '台'
    ws['H5'] = '旧备注'
    
    ws['A6'] = '5'
    ws['B6'] = 'OLD-005'
    ws['C6'] = '西门子'
    ws['D6'] = '模块'
    ws['E6'] = 'OLD-MODULE'
    ws['F6'] = '1'
    ws['G6'] = '个'
    ws['H6'] = '旧备注'
    
    wb.save(filename)
    wb.close()
    print(f"Created test Excel: {filename}")

# Test with res.md
print("=" * 60)
print("Testing with res.md (without ERP识别码 column)")
print("=" * 60)

test_file_res = 'test_excel_res.xlsx'
create_test_excel(test_file_res)

result_res = modify_excel(test_file_res, res_content, 'test_output')
print(f"\nResult: {result_res}")

if result_res['success']:
    print(f"✓ SUCCESS: {result_res['filename']}")
    
    # Verify the output
    from openpyxl import load_workbook
    wb_verify = load_workbook(result_res['output_path'])
    ws_verify = wb_verify.active
    
    print(f"\nVerifying output:")
    print(f"  Row 1: {[ws_verify.cell(1, i).value for i in range(1, 9)]}")
    print(f"  Row 3: {ws_verify['A3'].value}")
    print(f"  Row 4 (headers): {[ws_verify.cell(4, i).value for i in range(1, 9)]}")
    print(f"  Row 5 (data): {[ws_verify.cell(5, i).value for i in range(1, 9)]}")
    print(f"  Row 6 (data): {[ws_verify.cell(6, i).value for i in range(1, 9)]}")
    
    wb_verify.close()
else:
    print(f"✗ FAILED: {result_res['error']}")

# Cleanup
if os.path.exists(test_file_res):
    os.remove(test_file_res)

# Test with res2.md
print("\n" + "=" * 60)
print("Testing with res2.md (with ERP识别码 column)")
print("=" * 60)

test_file_res2 = 'test_excel_res2.xlsx'
create_test_excel(test_file_res2)

result_res2 = modify_excel(test_file_res2, res2_content, 'test_output')
print(f"\nResult: {result_res2}")

if result_res2['success']:
    print(f"✓ SUCCESS: {result_res2['filename']}")
    
    # Verify the output
    wb_verify2 = load_workbook(result_res2['output_path'])
    ws_verify2 = wb_verify2.active
    
    print(f"\nVerifying output:")
    print(f"  Row 1: {[ws_verify2.cell(1, i).value for i in range(1, 9)]}")
    print(f"  Row 3: {ws_verify2['A3'].value}")
    print(f"  Row 4 (headers): {[ws_verify2.cell(4, i).value for i in range(1, 9)]}")
    print(f"  Row 5 (data): {[ws_verify2.cell(5, i).value for i in range(1, 9)]}")
    print(f"  Total rows: {ws_verify2.max_row}")
    
    wb_verify2.close()
else:
    print(f"✗ FAILED: {result_res2['error']}")

# Cleanup
if os.path.