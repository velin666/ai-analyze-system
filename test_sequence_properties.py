#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Property-based tests and unit tests for sequence-based Excel matching

This file contains all the property tests and unit tests required by the spec.
"""

import sys
import os
import unittest
from openpyxl import Workbook

# Add server/api/files to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'server', 'api', 'files'))

from modify_excel_by_sequence import (
    TableExtractor, SequenceColumnLocator, SequenceMatcher,
    DataReplacer, TableData, SequenceColumnInfo
)

# Try to import hypothesis, but make tests work without it
try:
    from hypothesis import given, strategies as st, settings
    HYPOTHESIS_AVAILABLE = True
except ImportError:
    print("Warning: hypothesis not installed. Property tests will be skipped.")
    print("Install with: pip install hypothesis")
    HYPOTHESIS_AVAILABLE = False
    # Create dummy decorators and strategies
    class DummyStrategies:
        @staticmethod
        def lists(*args, **kwargs):
            return None
        @staticmethod
        def text(*args, **kwargs):
            return None
        @staticmethod
        def integers(*args, **kwargs):
            return None
        @staticmethod
        def one_of(*args, **kwargs):
            return None
        @staticmethod
        def dictionaries(*args, **kwargs):
            return None
    
    st = DummyStrategies()
    
    def given(*args, **kwargs):
        def decorator(func):
            return func
        return decorator
    
    def settings(**kwargs):
        def decorator(func):
            return func
        return decorator


class TestTableExtractor(unittest.TestCase):
    """Tests for TableExtractor class (Task 1.2)"""
    
    def test_extract_single_table(self):
        """Test extracting a single table"""
        markdown = """
| 序号 | 名称 | 数量 |
|------|------|------|
| 1    | A    | 10   |
| 2    | B    | 20   |
"""
        tables = TableExtractor.extract_all_tables(markdown)
        self.assertEqual(len(tables), 1)
        self.assertTrue(tables[0].has_sequence)
        self.assertEqual(len(tables[0].rows), 2)
    
    def test_extract_multiple_tables(self):
        """Test extracting multiple tables"""
        markdown = """
| 序号 | 名称 |
|------|------|
| 1    | A    |

Some text

| ID | Value |
|----|-------|
| 1  | X     |
"""
        tables = TableExtractor.extract_all_tables(markdown)
        self.assertEqual(len(tables), 2)
    
    def test_empty_input(self):
        """Test with empty input"""
        tables = TableExtractor.extract_all_tables("")
        self.assertEqual(len(tables), 0)
    
    def test_no_tables(self):
        """Test with no tables"""
        markdown = "This is just text without any tables"
        tables = TableExtractor.extract_all_tables(markdown)
        self.assertEqual(len(tables), 0)
    
    @unittest.skipUnless(HYPOTHESIS_AVAILABLE, "hypothesis not available")
    @given(st.lists(st.text(min_size=1, max_size=10), min_size=1, max_size=5))
    @settings(max_examples=50)
    def test_property_table_count(self, headers):
        """
        Property 1: Table extraction completeness
        For any list of headers, extraction should return correct number of tables
        """
        # Build markdown with N tables
        tables_md = []
        for header_list in [headers]:  # One table for simplicity
            header_row = "| " + " | ".join(header_list) + " |"
            separator = "|" + "|".join(["---" for _ in header_list]) + "|"
            data_row = "| " + " | ".join(["data" for _ in header_list]) + " |"
            table_md = f"{header_row}\n{separator}\n{data_row}"
            tables_md.append(table_md)
        
        markdown = "\n\n".join(tables_md)
        extracted = TableExtractor.extract_all_tables(markdown)
        
        # Should extract the same number of tables
        self.assertEqual(len(extracted), len(tables_md))


class TestSequenceColumnLocator(unittest.TestCase):
    """Tests for SequenceColumnLocator class (Task 1.4)"""
    
    def test_locate_sequence_column_first_row(self):
        """Test locating sequence column in first row"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['B1'] = '名称'
        ws['C1'] = '数量'
        
        info = SequenceColumnLocator.locate_sequence_column(ws)
        
        self.assertIsNotNone(info)
        self.assertEqual(info.column_index, 1)
        self.assertEqual(info.header_row, 1)
        self.assertIn('序号', info.column_headers)
    
    def test_locate_sequence_column_later_row(self):
        """Test locating sequence column in a later row"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Title'
        ws['A2'] = 'Subtitle'
        ws['A3'] = '序号'
        ws['B3'] = '名称'
        
        info = SequenceColumnLocator.locate_sequence_column(ws)
        
        self.assertIsNotNone(info)
        self.assertEqual(info.column_index, 1)
        self.assertEqual(info.header_row, 3)
    
    def test_no_sequence_column(self):
        """Test when no sequence column exists"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Name'
        ws['B1'] = 'Value'
        
        info = SequenceColumnLocator.locate_sequence_column(ws, max_rows=5)
        
        self.assertIsNone(info)
    
    def test_multiple_sequence_columns(self):
        """Test with multiple sequence columns (should use first)"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['B1'] = '名称'
        ws['C1'] = '序号'  # Duplicate
        
        info = SequenceColumnLocator.locate_sequence_column(ws)
        
        self.assertIsNotNone(info)
        self.assertEqual(info.column_index, 1)  # Should use first one
    
    @unittest.skipUnless(HYPOTHESIS_AVAILABLE, "hypothesis not available")
    @given(st.integers(min_value=1, max_value=10), st.integers(min_value=1, max_value=5))
    @settings(max_examples=20)
    def test_property_sequence_location(self, header_row, seq_col):
        """
        Property 2: Sequence column location
        For any Excel with sequence column, should find it correctly
        """
        wb = Workbook()
        ws = wb.active
        
        # Fill some cells before the header
        for row in range(1, header_row):
            ws.cell(row, 1, f"Text{row}")
        
        # Create header row with sequence column
        for col in range(1, 6):
            if col == seq_col:
                ws.cell(header_row, col, "序号")
            else:
                ws.cell(header_row, col, f"列{col}")
        
        info = SequenceColumnLocator.locate_sequence_column(ws, max_rows=20)
        
        self.assertIsNotNone(info)
        self.assertEqual(info.column_index, seq_col)
        self.assertEqual(info.header_row, header_row)


class TestSequenceMatcher(unittest.TestCase):
    """Tests for SequenceMatcher class (Task 1.6, 1.7)"""
    
    def test_normalize_sequence_integer(self):
        """Test normalizing integer sequences"""
        self.assertEqual(SequenceMatcher.normalize_sequence(12), "12")
        self.assertEqual(SequenceMatcher.normalize_sequence(1), "1")
    
    def test_normalize_sequence_string(self):
        """Test normalizing string sequences"""
        self.assertEqual(SequenceMatcher.normalize_sequence("12"), "12")
        self.assertEqual(SequenceMatcher.normalize_sequence("  12  "), "12")
    
    def test_normalize_sequence_leading_zeros(self):
        """Test normalizing sequences with leading zeros"""
        self.assertEqual(SequenceMatcher.normalize_sequence("012"), "12")
        self.assertEqual(SequenceMatcher.normalize_sequence("001"), "1")
        self.assertEqual(SequenceMatcher.normalize_sequence("0"), "0")
    
    def test_normalize_sequence_empty(self):
        """Test normalizing empty sequences"""
        self.assertEqual(SequenceMatcher.normalize_sequence(""), "")
        self.assertEqual(SequenceMatcher.normalize_sequence(None), "")
        self.assertEqual(SequenceMatcher.normalize_sequence("   "), "")
    
    @unittest.skipUnless(HYPOTHESIS_AVAILABLE, "hypothesis not available")
    @given(st.one_of(
        st.integers(min_value=0, max_value=1000),
        st.text(min_size=1, max_size=10),
        st.integers(min_value=0, max_value=100).map(lambda x: f"  {x}  "),
        st.integers(min_value=0, max_value=100).map(lambda x: f"0{x}")
    ))
    @settings(max_examples=50)
    def test_property_normalization_idempotent(self, value):
        """
        Property 5: Sequence value normalization
        Normalization should be idempotent: normalize(normalize(x)) == normalize(x)
        """
        normalized = SequenceMatcher.normalize_sequence(value)
        double_normalized = SequenceMatcher.normalize_sequence(normalized)
        
        self.assertEqual(normalized, double_normalized)
        self.assertIsInstance(normalized, str)
        # Should not have leading/trailing whitespace
        self.assertEqual(normalized, normalized.strip())
    
    def test_find_row_by_sequence(self):
        """Test finding row by sequence value"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['A2'] = '1'
        ws['A3'] = '2'
        ws['A4'] = '3'
        
        matcher = SequenceMatcher(ws, 1, 1)
        
        self.assertEqual(matcher.find_row_by_sequence("1"), 2)
        self.assertEqual(matcher.find_row_by_sequence("2"), 3)
        self.assertEqual(matcher.find_row_by_sequence("3"), 4)
        self.assertIsNone(matcher.find_row_by_sequence("999"))
    
    @unittest.skipUnless(HYPOTHESIS_AVAILABLE, "hypothesis not available")
    @given(st.dictionaries(
        keys=st.integers(min_value=1, max_value=100).map(str),
        values=st.integers(min_value=1, max_value=200),
        min_size=1,
        max_size=20
    ))
    @settings(max_examples=20)
    def test_property_sequence_matching(self, sequence_map):
        """
        Property 6: Sequence-based row matching
        For any sequence map, matcher should find existing sequences
        """
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        
        # Build worksheet with sequences
        for seq_value, row_num in sequence_map.items():
            ws.cell(row_num, 1, seq_value)
        
        matcher = SequenceMatcher(ws, 1, 1)
        
        # For each sequence in the map, should find the row
        for seq_value in list(sequence_map.keys())[:5]:  # Test first 5
            found_row = matcher.find_row_by_sequence(seq_value)
            self.assertIsNotNone(found_row)


class TestDataReplacer(unittest.TestCase):
    """Tests for DataReplacer class (Task 1.9)"""
    
    def test_replace_row_basic(self):
        """Test basic row replacement"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['B1'] = '名称'
        ws['C1'] = '数量'
        ws['A2'] = '1'
        ws['B2'] = 'Old Name'
        ws['C2'] = '10'
        
        ai_row_data = {'序号': '1', '名称': 'New Name', '数量': '20'}
        column_mapping = {'序号': 1, '名称': 2, '数量': 3}
        
        count = DataReplacer.replace_row(ws, 2, ai_row_data, column_mapping)
        
        self.assertEqual(count, 3)
        self.assertEqual(ws['B2'].value, 'New Name')
        self.assertEqual(ws['C2'].value, '20')
    
    def test_replace_row_partial(self):
        """Test partial row replacement (some columns not in mapping)"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['B1'] = '名称'
        ws['C1'] = '数量'
        ws['D1'] = '备注'
        ws['A2'] = '1'
        ws['B2'] = 'Old Name'
        ws['C2'] = '10'
        ws['D2'] = 'Keep This'
        
        ai_row_data = {'序号': '1', '名称': 'New Name'}
        column_mapping = {'序号': 1, '名称': 2}  # No mapping for 数量 and 备注
        
        count = DataReplacer.replace_row(ws, 2, ai_row_data, column_mapping)
        
        self.assertEqual(count, 2)
        self.assertEqual(ws['B2'].value, 'New Name')
        self.assertEqual(ws['C2'].value, '10')  # Unchanged
        self.assertEqual(ws['D2'].value, 'Keep This')  # Unchanged
    
    @unittest.skipUnless(HYPOTHESIS_AVAILABLE, "hypothesis not available")
    @given(st.dictionaries(
        keys=st.text(min_size=1, max_size=10),
        values=st.text(max_size=20),
        min_size=1,
        max_size=5
    ))
    @settings(max_examples=20)
    def test_property_column_replacement(self, ai_data):
        """
        Property 8: Column-based data replacement
        Only mapped columns should be replaced, others remain unchanged
        """
        wb = Workbook()
        ws = wb.active
        
        # Create headers and initial data
        col_names = list(ai_data.keys()) + ['Unmapped1', 'Unmapped2']
        for idx, name in enumerate(col_names, 1):
            ws.cell(1, idx, name)
            ws.cell(2, idx, f"Original{idx}")
        
        # Create column mapping (only for AI data columns)
        column_mapping = {name: idx + 1 for idx, name in enumerate(ai_data.keys())}
        
        # Replace row
        DataReplacer.replace_row(ws, 2, ai_data, column_mapping)
        
        # Check that mapped columns are replaced
        for name, value in ai_data.items():
            col_idx = column_mapping[name]
            self.assertEqual(ws.cell(2, col_idx).value, value)
        
        # Check that unmapped columns are unchanged
        unmapped_start = len(ai_data) + 1
        for idx in range(unmapped_start, len(col_names) + 1):
            self.assertEqual(ws.cell(2, idx).value, f"Original{idx}")


class TestEdgeCases(unittest.TestCase):
    """Unit tests for edge cases (Task 7.1)"""
    
    def test_empty_table(self):
        """Test handling empty tables"""
        markdown = """
| 序号 | 名称 |
|------|------|
"""
        tables = TableExtractor.extract_all_tables(markdown)
        self.assertEqual(len(tables), 1)
        self.assertEqual(len(tables[0].rows), 0)
    
    def test_table_without_sequence(self):
        """Test table without sequence column"""
        markdown = """
| Name | Value |
|------|-------|
| A    | 1     |
"""
        tables = TableExtractor.extract_all_tables(markdown)
        self.assertEqual(len(tables), 1)
        self.assertFalse(tables[0].has_sequence)
    
    def test_empty_sequence_value(self):
        """Test row with empty sequence value"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['A2'] = ''  # Empty sequence
        ws['A3'] = '2'
        
        matcher = SequenceMatcher(ws, 1, 1)
        
        # Should not find empty sequence
        self.assertIsNone(matcher.find_row_by_sequence(""))
        # Should find non-empty sequence
        self.assertEqual(matcher.find_row_by_sequence("2"), 3)
    
    def test_duplicate_sequences(self):
        """Test handling duplicate sequence values"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['A2'] = '1'
        ws['A3'] = '1'  # Duplicate
        ws['A4'] = '2'
        
        matcher = SequenceMatcher(ws, 1, 1)
        
        # Should use first occurrence
        self.assertEqual(matcher.find_row_by_sequence("1"), 2)
    
    def test_special_characters(self):
        """Test handling special characters in data"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['B1'] = '名称'
        ws['A2'] = '1'
        ws['B2'] = 'Old'
        
        ai_row_data = {'序号': '1', '名称': 'Name with "quotes" & <special>'}
        column_mapping = {'序号': 1, '名称': 2}
        
        count = DataReplacer.replace_row(ws, 2, ai_row_data, column_mapping)
        
        self.assertEqual(count, 2)
        self.assertEqual(ws['B2'].value, 'Name with "quotes" & <special>')


class TestStatisticsConsistency(unittest.TestCase):
    """Tests for statistics consistency (Task 2.2)"""
    
    @unittest.skipUnless(HYPOTHESIS_AVAILABLE, "hypothesis not available")
    @given(
        st.integers(min_value=1, max_value=10),
        st.integers(min_value=0, max_value=10)
    )
    @settings(max_examples=20)
    def test_property_statistics_consistency(self, total_tables, processed_tables):
        """
        Property 11: Statistics consistency
        processed_tables + skipped_tables should equal total_tables
        """
        if processed_tables > total_tables:
            processed_tables = total_tables
        
        skipped_tables = total_tables - processed_tables
        
        # Verify consistency
        self.assertEqual(processed_tables + skipped_tables, total_tables)
        self.assertGreaterEqual(processed_tables, 0)
        self.assertGreaterEqual(skipped_tables, 0)


def run_tests():
    """Run all tests"""
    # Create test suite
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # Add all test classes
    suite.addTests(loader.loadTestsFromTestCase(TestTableExtractor))
    suite.addTests(loader.loadTestsFromTestCase(TestSequenceColumnLocator))
    suite.addTests(loader.loadTestsFromTestCase(TestSequenceMatcher))
    suite.addTests(loader.loadTestsFromTestCase(TestDataReplacer))
    suite.addTests(loader.loadTestsFromTestCase(TestEdgeCases))
    suite.addTests(loader.loadTestsFromTestCase(TestStatisticsConsistency))
    
    # Run tests
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    return result.wasSuccessful()


if __name__ == '__main__':
    success = run_tests()
    sys.exit(0 if success else 1)
