#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
验证Excel输出文件的正确性
"""

from openpyxl import load_workbook

def verify_output():
    """验证输出文件"""
    output_path = 'uploads/modified/KHG51-SD01 烘烤炉电气件清单(修改后)_7.xlsx'
    
    print("=" * 80)
    print("验证Excel输出文件")
    print("=" * 80)
    
    wb = load_workbook(output_path)
    ws = wb.active
    
    # 找到序号列和表头行
    seq_col = 1
    header_row = 4
    
    # 验证关键行的数据
    test_cases = [
        (1, 5, "序号1应该是HMI面板"),
        (3, 7, "序号3应该是SMART模拟量输入模块"),
        (12, 16, "序号12应该是变频器0.75KW"),
        (27, 31, "序号27应该是中间小型继电器21只"),
        (28, 32, "序号28应该是中间小型继电器9只"),
    ]
    
    print("\n验证关键行数据：")
    print("-" * 80)
    
    for seq_num, row_num, description in test_cases:
        seq_value = ws.cell(row_num, seq_col).value
        name_value = ws.cell(row_num, 4).value  # 名称列
        model_value = ws.cell(row_num, 5).value  # 型号列
        qty_value = ws.cell(row_num, 6).value  # 数量列
        
        print(f"\n序号 {seq_num} (Excel第{row_num}行):")
        print(f"  序号值: {seq_value}")
        print(f"  名称: {name_value}")
        print(f"  型号: {model_value}")
        print(f"  数量: {qty_value}")
        print(f"  预期: {description}")
        
        # 验证序号是否正确
        if str(seq_value) == str(seq_num):
            print(f"  ✓ 序号正确")
        else:
            print(f"  ✗ 序号错误！应该是{seq_num}，实际是{seq_value}")
    
    print("\n" + "=" * 80)
    print("验证完成")
    print("=" * 80)

if __name__ == '__main__':
    verify_output()
