#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
快速验证Excel修改功能
运行此脚本可以快速验证Excel修改功能是否正常工作
"""

import sys
import os
sys.path.insert(0, 'server/api/files')

from modify_excel_by_sequence import modify_excel_by_sequence
import openpyxl

def quick_test():
    """快速测试Excel修改功能"""
    
    print("="*60)
    print("Excel修改功能快速验证")
    print("="*60)
    
    # 测试res4.md（最容易出问题的场景）
    print("\n测试场景: res4.md - 部分行修改")
    print("-"*60)
    
    # 读取res4.md
    with open('server/api/res4.md', 'r', encoding='utf-8') as f:
        ai_result = f.read()
    
    # 执行修改
    result = modify_excel_by_sequence(
        'assets/KHG51-SD01 烘烤炉电气件清单.xlsx',
        ai_result,
        'uploads/modified'
    )
    
    if not result['success']:
        print(f"✗ 测试失败: {result.get('error')}")
        return False
    
    print(f"✓ 处理成功: {result['statistics']['matched_rows']}行匹配")
    
    # 验证关键数据
    print("\n验证关键数据:")
    print("-"*60)
    
    wb = openpyxl.load_workbook(result['output_path'])
    ws = wb.active
    
    # 验证序号1（不应该被修改）
    seq1_brand = ws.cell(5, 3).value  # 第5行，品牌列
    if 'SIEMENS' in str(seq1_brand) or '西门子' in str(seq1_brand):
        print("✓ 序号1: 保持原样（西门子）- 正确")
    else:
        print(f"✗ 序号1: 品牌被错误修改为 {seq1_brand}")
        wb.close()
        return False
    
    # 验证序号6（应该被修改为三菱）
    seq6_brand = ws.cell(10, 3).value  # 第10行，品牌列
    seq6_name = ws.cell(10, 4).value   # 第10行，名称列
    if '三菱' in str(seq6_brand) and '模拟量' in str(seq6_name):
        print("✓ 序号6: 正确修改为三菱模拟量模块 - 正确")
    else:
        print(f"✗ 序号6: 品牌={seq6_brand}, 名称={seq6_name} - 错误")
        wb.close()
        return False
    
    # 验证序号107（应该被正确修改）
    seq107_name = ws.cell(111, 4).value  # 第111行，名称列
    if '电气配件' in str(seq107_name):
        print("✓ 序号107: 正确修改为电气配件 - 正确")
    else:
        print(f"✗ 序号107: 名称={seq107_name} - 错误")
        wb.close()
        return False
    
    wb.close()
    
    print("\n" + "="*60)
    print("✓ 所有验证通过！Excel修改功能工作正常。")
    print("="*60)
    
    return True

if __name__ == '__main__':
    try:
        success = quick_test()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n✗ 测试过程中发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
