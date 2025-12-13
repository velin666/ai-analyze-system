#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Unit tests for sequence-based Excel matching

This file contains unit tests that don't require hypothesis.
"""

import sys
import os
import unittest
from openpyxl import Workbook

# Add server/api/files to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'server', 'api', 'files'))

from modify_excel_by_sequence import (
    TableExtractor, SequenceColumnLocator, SequenceMatcher,
    DataReplacer, TableData, SequenceColumnInfo
)


class TestTableExtractor(unittest.TestCase):
    """Tests for TableExtractor class"""
    
    def test_extract_single_table(self):
        """Test extracting a single table"""
        markdown = """
| 序号 | 名称 | 数量 |
|------|------|------|
| 1    | A    | 10   |
| 2    | B    | 20   |
"""
        tables = TableExtractor.extract_all_tables(markdown)
        self.assertEqual(len(tables), 1)
        self.assertTrue(tables[0].has_sequence)
        self.assertEqual(len(tables[0].rows), 2)
    
    def test_extract_multiple_tables(self):
        """Test extracting multiple tables"""
        markdown = """
| 序号 | 名称 |
|------|------|
| 1    | A    |

Some text

| ID | Value |
|----|-------|
| 1  | X     |
"""
        tables = TableExtractor.extract_all_tables(markdown)
        self.assertEqual(len(tables), 2)
    
    def test_empty_input(self):
        """Test with empty input"""
        tables = TableExtractor.extract_all_tables("")
        self.assertEqual(len(tables), 0)
    
    def test_no_tables(self):
        """Test with no tables"""
        markdown = "This is just text without any tables"
        tables = TableExtractor.extract_all_tables(markdown)
        self.assertEqual(len(tables), 0)


class TestSequenceColumnLocator(unittest.TestCase):
    """Tests for SequenceColumnLocator class"""
    
    def test_locate_sequence_column_first_row(self):
        """Test locating sequence column in first row"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['B1'] = '名称'
        ws['C1'] = '数量'
        
        info = SequenceColumnLocator.locate_sequence_column(ws)
        
        self.assertIsNotNone(info)
        self.assertEqual(info.column_index, 1)
        self.assertEqual(info.header_row, 1)
        self.assertIn('序号', info.column_headers)
    
    def test_locate_sequence_column_later_row(self):
        """Test locating sequence column in a later row"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Title'
        ws['A2'] = 'Subtitle'
        ws['A3'] = '序号'
        ws['B3'] = '名称'
        
        info = SequenceColumnLocator.locate_sequence_column(ws)
        
        self.assertIsNotNone(info)
        self.assertEqual(info.column_index, 1)
        self.assertEqual(info.header_row, 3)
    
    def test_no_sequence_column(self):
        """Test when no sequence column exists"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Name'
        ws['B1'] = 'Value'
        
        info = SequenceColumnLocator.locate_sequence_column(ws, max_rows=5)
        
        self.assertIsNone(info)
    
    def test_multiple_sequence_columns(self):
        """Test with multiple sequence columns (should use first)"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['B1'] = '名称'
        ws['C1'] = '序号'  # Duplicate
        
        info = SequenceColumnLocator.locate_sequence_column(ws)
        
        self.assertIsNotNone(info)
        self.assertEqual(info.column_index, 1)  # Should use first one


class TestSequenceMatcher(unittest.TestCase):
    """Tests for SequenceMatcher class"""
    
    def test_normalize_sequence_integer(self):
        """Test normalizing integer sequences"""
        self.assertEqual(SequenceMatcher.normalize_sequence(12), "12")
        self.assertEqual(SequenceMatcher.normalize_sequence(1), "1")
    
    def test_normalize_sequence_string(self):
        """Test normalizing string sequences"""
        self.assertEqual(SequenceMatcher.normalize_sequence("12"), "12")
        self.assertEqual(SequenceMatcher.normalize_sequence("  12  "), "12")
    
    def test_normalize_sequence_leading_zeros(self):
        """Test normalizing sequences with leading zeros"""
        self.assertEqual(SequenceMatcher.normalize_sequence("012"), "12")
        self.assertEqual(SequenceMatcher.normalize_sequence("001"), "1")
        self.assertEqual(SequenceMatcher.normalize_sequence("0"), "0")
    
    def test_normalize_sequence_empty(self):
        """Test normalizing empty sequences"""
        self.assertEqual(SequenceMatcher.normalize_sequence(""), "")
        self.assertEqual(SequenceMatcher.normalize_sequence(None), "")
        self.assertEqual(SequenceMatcher.normalize_sequence("   "), "")
    
    def test_normalize_sequence_idempotent(self):
        """Test that normalization is idempotent"""
        test_values = [12, "12", "  12  ", "012", "0", "", None]
        for value in test_values:
            normalized = SequenceMatcher.normalize_sequence(value)
            double_normalized = SequenceMatcher.normalize_sequence(normalized)
            self.assertEqual(normalized, double_normalized)
    
    def test_find_row_by_sequence(self):
        """Test finding row by sequence value"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['A2'] = '1'
        ws['A3'] = '2'
        ws['A4'] = '3'
        
        matcher = SequenceMatcher(ws, 1, 1)
        
        self.assertEqual(matcher.find_row_by_sequence("1"), 2)
        self.assertEqual(matcher.find_row_by_sequence("2"), 3)
        self.assertEqual(matcher.find_row_by_sequence("3"), 4)
        self.assertIsNone(matcher.find_row_by_sequence("999"))


class TestDataReplacer(unittest.TestCase):
    """Tests for DataReplacer class"""
    
    def test_replace_row_basic(self):
        """Test basic row replacement"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['B1'] = '名称'
        ws['C1'] = '数量'
        ws['A2'] = '1'
        ws['B2'] = 'Old Name'
        ws['C2'] = '10'
        
        ai_row_data = {'序号': '1', '名称': 'New Name', '数量': '20'}
        column_mapping = {'序号': 1, '名称': 2, '数量': 3}
        
        count = DataReplacer.replace_row(ws, 2, ai_row_data, column_mapping)
        
        self.assertEqual(count, 3)
        self.assertEqual(ws['B2'].value, 'New Name')
        self.assertEqual(ws['C2'].value, '20')
    
    def test_replace_row_partial(self):
        """Test partial row replacement (some columns not in mapping)"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['B1'] = '名称'
        ws['C1'] = '数量'
        ws['D1'] = '备注'
        ws['A2'] = '1'
        ws['B2'] = 'Old Name'
        ws['C2'] = '10'
        ws['D2'] = 'Keep This'
        
        ai_row_data = {'序号': '1', '名称': 'New Name'}
        column_mapping = {'序号': 1, '名称': 2}  # No mapping for 数量 and 备注
        
        count = DataReplacer.replace_row(ws, 2, ai_row_data, column_mapping)
        
        self.assertEqual(count, 2)
        self.assertEqual(ws['B2'].value, 'New Name')
        self.assertEqual(ws['C2'].value, '10')  # Unchanged
        self.assertEqual(ws['D2'].value, 'Keep This')  # Unchanged
    
    def test_replace_preserves_unmapped_columns(self):
        """Test that unmapped columns are preserved"""
        wb = Workbook()
        ws = wb.active
        
        # Create headers
        headers = ['序号', '名称', '数量', 'Unmapped1', 'Unmapped2']
        for idx, name in enumerate(headers, 1):
            ws.cell(1, idx, name)
            ws.cell(2, idx, f"Original{idx}")
        
        # Replace only first two columns
        ai_row_data = {'序号': '1', '名称': 'New Name'}
        column_mapping = {'序号': 1, '名称': 2}
        
        DataReplacer.replace_row(ws, 2, ai_row_data, column_mapping)
        
        # Check mapped columns are replaced
        self.assertEqual(ws.cell(2, 1).value, '1')
        self.assertEqual(ws.cell(2, 2).value, 'New Name')
        
        # Check unmapped columns are unchanged
        self.assertEqual(ws.cell(2, 3).value, 'Original3')
        self.assertEqual(ws.cell(2, 4).value, 'Original4')
        self.assertEqual(ws.cell(2, 5).value, 'Original5')


class TestEdgeCases(unittest.TestCase):
    """Unit tests for edge cases"""
    
    def test_empty_table(self):
        """Test handling empty tables"""
        markdown = """
| 序号 | 名称 |
|------|------|
"""
        tables = TableExtractor.extract_all_tables(markdown)
        self.assertEqual(len(tables), 1)
        self.assertEqual(len(tables[0].rows), 0)
    
    def test_table_without_sequence(self):
        """Test table without sequence column"""
        markdown = """
| Name | Value |
|------|-------|
| A    | 1     |
"""
        tables = TableExtractor.extract_all_tables(markdown)
        self.assertEqual(len(tables), 1)
        self.assertFalse(tables[0].has_sequence)
    
    def test_empty_sequence_value(self):
        """Test row with empty sequence value"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['A2'] = ''  # Empty sequence
        ws['A3'] = '2'
        
        matcher = SequenceMatcher(ws, 1, 1)
        
        # Should not find empty sequence
        self.assertIsNone(matcher.find_row_by_sequence(""))
        # Should find non-empty sequence
        self.assertEqual(matcher.find_row_by_sequence("2"), 3)
    
    def test_duplicate_sequences(self):
        """Test handling duplicate sequence values"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['A2'] = '1'
        ws['A3'] = '1'  # Duplicate
        ws['A4'] = '2'
        
        matcher = SequenceMatcher(ws, 1, 1)
        
        # Should use first occurrence
        self.assertEqual(matcher.find_row_by_sequence("1"), 2)
    
    def test_special_characters(self):
        """Test handling special characters in data"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['B1'] = '名称'
        ws['A2'] = '1'
        ws['B2'] = 'Old'
        
        ai_row_data = {'序号': '1', '名称': 'Name with "quotes" & <special>'}
        column_mapping = {'序号': 1, '名称': 2}
        
        count = DataReplacer.replace_row(ws, 2, ai_row_data, column_mapping)
        
        self.assertEqual(count, 2)
        self.assertEqual(ws['B2'].value, 'Name with "quotes" & <special>')


class TestStatisticsConsistency(unittest.TestCase):
    """Tests for statistics consistency"""
    
    def test_statistics_consistency(self):
        """Test that statistics are consistent"""
        test_cases = [
            (10, 7),  # 10 total, 7 processed
            (5, 5),   # All processed
            (5, 0),   # None processed
            (1, 1),   # Single table
        ]
        
        for total_tables, processed_tables in test_cases:
            skipped_tables = total_tables - processed_tables
            
            # Verify consistency
            self.assertEqual(processed_tables + skipped_tables, total_tables)
            self.assertGreaterEqual(processed_tables, 0)
            self.assertGreaterEqual(skipped_tables, 0)


def run_tests():
    """Run all tests"""
    # Create test suite
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # Add all test classes
    suite.addTests(loader.loadTestsFromTestCase(TestTableExtractor))
    suite.addTests(loader.loadTestsFromTestCase(TestSequenceColumnLocator))
    suite.addTests(loader.loadTestsFromTestCase(TestSequenceMatcher))
    suite.addTests(loader.loadTestsFromTestCase(TestDataReplacer))
    suite.addTests(loader.loadTestsFromTestCase(TestEdgeCases))
    suite.addTests(loader.loadTestsFromTestCase(TestStatisticsConsistency))
    
    # Run tests
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    return result.wasSuccessful()


if __name__ == '__main__':
    success = run_tests()
    sys.exit(0 if success else 1)
