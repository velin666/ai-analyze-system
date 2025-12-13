#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel基于序号列的匹配和替换脚本

实现基于"序号"列的简化匹配功能：
- 序号列定位：在Excel前20行中查找"序号"列
- 直接匹配：通过序号值直接定位目标行（O(1)查找）
- 简化逻辑：无需多列比对，避免误匹配
- 高性能：构建序号映射表，快速查找

版本: 1.0
作者: AI Assistant
日期: 2024
"""

import sys
import json
import os
import time
import logging
import re
from typing import List, Dict, Optional, Any, Tuple
from dataclasses import dataclass, field
from openpyxl import load_workbook

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# 数据模型类 (Task 1.10)
# ============================================================================

@dataclass
class TableData:
    """表格数据结构"""
    headers: List[str]                      # 表头列表
    rows: List[Dict[str, Any]]              # 数据行列表（列名到值的字典）
    has_sequence: bool                      # 是否包含序号列
    sequence_col_index: int = -1            # 序号列在AI表格中的索引（-1表示不存在）


@dataclass
class SequenceColumnInfo:
    """序号列信息"""
    column_index: int                       # 序号列的列索引（从1开始）
    header_row: int                         # 表头所在行号
    column_headers: Dict[str, int]          # 所有列名到列索引的映射


@dataclass
class ProcessingStatistics:
    """处理统计信息"""
    total_tables: int = 0                   # 总表格数
    processed_tables: int = 0               # 成功处理的表格数
    skipped_tables: int = 0                 # 跳过的表格数（无序号列）
    total_rows: int = 0                     # 总行数
    matched_rows: int = 0                   # 成功匹配的行数
    skipped_rows: int = 0                   # 跳过的行数（序号未找到）
    processing_time: float = 0.0            # 处理时间（秒）


@dataclass
class ProcessingResult:
    """处理结果"""
    success: bool                           # 是否成功
    output_path: Optional[str] = None       # 输出文件路径
    filename: Optional[str] = None          # 输出文件名
    statistics: ProcessingStatistics = field(default_factory=ProcessingStatistics)
    error: Optional[str] = None             # 错误信息
    warnings: List[str] = field(default_factory=list)  # 警告列表


@dataclass
class ProcessingConfig:
    """处理配置"""
    max_header_search_rows: int = 20        # 最大表头搜索行数
    normalize_sequence: bool = True         # 是否标准化序号值
    skip_empty_sequence: bool = True        # 是否跳过空序号行
    log_level: str = "INFO"                 # 日志级别


# ============================================================================
# 自定义异常类
# ============================================================================

class SequenceColumnNotFoundError(Exception):
    """序号列未找到异常"""
    pass


# ============================================================================
# TableExtractor类 (Task 1.1)
# ============================================================================

class TableExtractor:
    """
    表格提取器 - 从Markdown文本中提取和解析表格
    
    负责从AI返回的Markdown格式文本中识别和提取所有表格，
    并将每个表格解析为结构化的TableData对象。
    
    与之前版本的区别：
    - 数据行转换为字典格式（列名到值的映射）
    - 检测表格是否包含"序号"列
    - 记录序号列的索引位置
    """
    
    @staticmethod
    def extract_all_tables(markdown_text: str) -> List[TableData]:
        """
        提取AI返回内容中的所有Markdown表格并解析
        
        Args:
            markdown_text: AI返回的Markdown格式文本
            
        Returns:
            TableData对象列表
        """
        if not markdown_text or not isinstance(markdown_text, str):
            logger.warning("AI结果为空或格式不正确")
            return []
        
        try:
            logger.info("开始从AI结果中提取表格...")
            
            # 处理转义的换行符
            if '\\n' in markdown_text:
                markdown_text = markdown_text.replace('\\n', '\n')
            
            tables = []
            lines = markdown_text.split('\n')
            current_table_lines = []
            in_table = False
            
            for line in lines:
                stripped = line.strip()
                
                # 检查是否是表格行（包含|符号）
                if '|' in stripped:
                    if not in_table:
                        in_table = True
                        current_table_lines = []
                    current_table_lines.append(line)
                else:
                    # 遇到非表格行，如果之前在表格中则表格结束
                    if in_table and current_table_lines:
                        table_data = TableExtractor.parse_table('\n'.join(current_table_lines))
                        if table_data:
                            tables.append(table_data)
                        current_table_lines = []
                        in_table = False
            
            # 处理文件末尾的表格
            if in_table and current_table_lines:
                table_data = TableExtractor.parse_table('\n'.join(current_table_lines))
                if table_data:
                    tables.append(table_data)
            
            if tables:
                logger.info(f"✓ 成功提取并解析 {len(tables)} 个表格")
                for idx, table in enumerate(tables, 1):
                    logger.info(f"  表格{idx}: {len(table.headers)}列, {len(table.rows)}行, "
                              f"包含序号列: {'是' if table.has_sequence else '否'}")
            else:
                logger.warning("未在AI结果中找到任何Markdown表格")
            
            return tables
            
        except Exception as e:
            logger.error(f"✗ 表格提取失败: {str(e)}", exc_info=True)
            return []
    
    @staticmethod
    def parse_table(table_text: str) -> Optional[TableData]:
        """
        解析单个Markdown表格为结构化数据
        
        Args:
            table_text: Markdown表格文本
            
        Returns:
            TableData对象，如果解析失败返回None
        """
        if not table_text:
            return None
        
        try:
            lines = [line.strip() for line in table_text.split('\n') if line.strip()]
            
            if len(lines) < 2:
                return None
            
            # 解析表头（第一行）
            header_line = lines[0]
            headers = [cell.strip() for cell in header_line.split('|')]
            headers = [h for h in headers if h]  # 移除空字符串
            
            if not headers:
                return None
            
            # 检查是否包含"序号"列
            has_sequence = False
            sequence_col_index = -1
            for idx, header in enumerate(headers):
                if header == "序号" or header.lower() == "序号":
                    has_sequence = True
                    sequence_col_index = idx
                    break
            
            # 解析数据行（从第三行开始，跳过分隔符行）
            rows = []
            has_ellipsis = False  # 标记是否包含省略号
            
            for line in lines[2:]:
                cells = [cell.strip() for cell in line.split('|')]
                cells = [c for c in cells if c or c == '']
                
                # 移除首尾的空字符串（由首尾的|产生）
                if cells and cells[0] == '':
                    cells = cells[1:]
                if cells and cells[-1] == '':
                    cells = cells[:-1]
                
                # 检查是否包含省略号（...）
                if any('...' in str(cell) for cell in cells):
                    has_ellipsis = True
                    logger.debug(f"  检测到省略号行，标记为示例表格")
                    continue  # 跳过包含省略号的行
                
                # 过滤完全空的行
                if cells and any(cell for cell in cells):
                    # 确保列数与表头一致
                    while len(cells) < len(headers):
                        cells.append('')
                    cells = cells[:len(headers)]
                    
                    # 转换为字典格式
                    row_dict = {headers[i]: cells[i] for i in range(len(headers))}
                    rows.append(row_dict)
            
            # 如果表格包含省略号，完全跳过该表格（即使有一些有效行）
            # 因为这通常是示例/摘要表格，不是完整数据
            if has_ellipsis:
                logger.warning(f"  ⚠ 检测到示例表格（包含省略号），完全跳过以避免数据覆盖")
                return None
            
            return TableData(
                headers=headers,
                rows=rows,
                has_sequence=has_sequence,
                sequence_col_index=sequence_col_index
            )
            
        except Exception as e:
            logger.error(f"✗ 表格解析失败: {str(e)}", exc_info=True)
            return None


# ============================================================================
# SequenceColumnLocator类 (Task 1.3)
# ============================================================================

class SequenceColumnLocator:
    """
    序号列定位器 - 在Excel文件中定位"序号"列
    
    负责在Excel文件的前N行中查找包含"序号"的表头行，
    并解析所有列名创建列映射。
    """
    
    @staticmethod
    def locate_sequence_column(worksheet, max_rows: int = 20) -> Optional[SequenceColumnInfo]:
        """
        在Excel工作表中定位序号列
        
        Args:
            worksheet: openpyxl工作表对象
            max_rows: 最大搜索行数，默认20
            
        Returns:
            SequenceColumnInfo对象，如果未找到返回None
        """
        try:
            logger.info(f"在前{max_rows}行中查找'序号'列...")
            
            max_scan_rows = min(max_rows, worksheet.max_row)
            
            for row_num in range(1, max_scan_rows + 1):
                row = worksheet[row_num]
                
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value is not None:
                        cell_value = str(cell.value).strip()
                        
                        # 检查是否是"序号"列
                        if cell_value == "序号" or cell_value.lower() == "序号":
                            logger.info(f"✓ 找到序号列: 第{row_num}行, 第{col_idx}列")
                            
                            # 解析该行的所有列名
                            column_headers = SequenceColumnLocator.parse_headers(worksheet, row_num)
                            
                            return SequenceColumnInfo(
                                column_index=col_idx,
                                header_row=row_num,
                                column_headers=column_headers
                            )
            
            logger.warning(f"✗ 在前{max_scan_rows}行中未找到'序号'列")
            return None
            
        except Exception as e:
            logger.error(f"✗ 定位序号列失败: {str(e)}", exc_info=True)
            return None
    
    @staticmethod
    def parse_headers(worksheet, header_row: int) -> Dict[str, int]:
        """
        解析表头行，创建列名到列索引的映射
        
        Args:
            worksheet: openpyxl工作表对象
            header_row: 表头所在行号
            
        Returns:
            列名到列索引的字典 {列名: 列索引(1-based)}
        """
        column_headers = {}
        row = worksheet[header_row]
        
        for col_idx, cell in enumerate(row, start=1):
            if cell.value is not None:
                # 去除所有空格（包括中间的空格）以实现更好的匹配
                header_name = str(cell.value).strip().replace(' ', '').replace('\u3000', '')
                if header_name:
                    column_headers[header_name] = col_idx
        
        logger.debug(f"  解析到{len(column_headers)}个列名: {list(column_headers.keys())}")
        return column_headers


# ============================================================================
# SequenceMatcher类 (Task 1.5)
# ============================================================================

class SequenceMatcher:
    """
    序号匹配器 - 基于序号值进行行匹配
    
    核心功能：
    - 构建序号到行号的映射表（O(1)查找）
    - 标准化序号值（处理前导零、空格、类型等）
    - 通过序号值直接查找目标行
    """
    
    def __init__(self, worksheet, sequence_col_index: int, header_row: int):
        """
        初始化序号匹配器
        
        Args:
            worksheet: openpyxl工作表对象
            sequence_col_index: 序号列的列索引（1-based）
            header_row: 表头所在行号
        """
        self.worksheet = worksheet
        self.sequence_col_index = sequence_col_index
        self.header_row = header_row
        self.sequence_map = self._build_sequence_map()
    
    def _build_sequence_map(self) -> Dict[str, int]:
        """
        构建序号到行号的映射表
        
        遍历序号列的所有数据行，创建序号值到行号的映射。
        使用标准化后的序号值作为键。
        
        Returns:
            序号值到行号的字典 {标准化序号值: 行号}
        """
        sequence_map = {}
        start_row = self.header_row + 1
        
        logger.info(f"构建序号映射表（从第{start_row}行开始）...")
        
        for row_num in range(start_row, self.worksheet.max_row + 1):
            cell = self.worksheet.cell(row_num, self.sequence_col_index)
            
            if cell.value is not None:
                normalized_seq = self.normalize_sequence(cell.value)
                
                if normalized_seq:  # 跳过空序号
                    if normalized_seq in sequence_map:
                        logger.warning(f"  发现重复序号 '{normalized_seq}' (行{sequence_map[normalized_seq]} 和 行{row_num}), 使用第一个")
                    else:
                        sequence_map[normalized_seq] = row_num
        
        logger.info(f"✓ 序号映射表构建完成，共{len(sequence_map)}个序号")
        return sequence_map
    
    def find_row_by_sequence(self, sequence_value: Any) -> Optional[int]:
        """
        通过序号值查找行号
        
        Args:
            sequence_value: 序号值（任意类型）
            
        Returns:
            匹配的行号，如果未找到返回None
        """
        normalized_seq = self.normalize_sequence(sequence_value)
        
        if not normalized_seq:
            return None
        
        return self.sequence_map.get(normalized_seq)
    
    @staticmethod
    def normalize_sequence(value: Any) -> str:
        """
        标准化序号值
        
        标准化规则：
        1. 转换为字符串
        2. 去除前导和尾随空格
        3. 去除前导零（但保留单个"0"）
        4. 统一处理None和空字符串
        
        Args:
            value: 原始序号值（可以是数字、字符串等）
            
        Returns:
            标准化后的字符串
        """
        if value is None or value == "":
            return ""
        
        # 转换为字符串并去除空格
        str_value = str(value).strip()
        
        if not str_value:
            return ""
        
        # 去除前导零，但保留单个"0"
        if str_value.isdigit():
            str_value = str(int(str_value))
        
        return str_value


# ============================================================================
# DataReplacer类 (Task 1.8)
# ============================================================================

class DataReplacer:
    """
    数据替换器 - 替换Excel中的行数据
    
    负责将AI数据行的内容精确替换到Excel文件的目标行。
    """
    
    @staticmethod
    def replace_row(worksheet,
                   row_number: int,
                   ai_row_data: Dict[str, Any],
                   column_mapping: Dict[str, int]) -> int:
        """
        替换指定行的数据
        
        Args:
            worksheet: openpyxl工作表对象
            row_number: 目标行号（1-based）
            ai_row_data: AI数据行（列名到值的字典）
            column_mapping: 列名到Excel列索引的映射
            
        Returns:
            实际替换的列数
        """
        try:
            replaced_count = 0
            
            for col_name, col_value in ai_row_data.items():
                # 检查该列是否在Excel中存在
                if col_name in column_mapping:
                    excel_col_idx = column_mapping[col_name]
                    cell = worksheet.cell(row_number, excel_col_idx)
                    
                    # 处理空值
                    if col_value is None or col_value == '':
                        new_value = ''
                    else:
                        new_value = str(col_value).strip()
                    
                    old_value = cell.value
                    cell.value = new_value
                    replaced_count += 1
                    
                    if old_value != new_value:
                        logger.debug(f"    列'{col_name}': '{old_value}' -> '{new_value}'")
            
            logger.debug(f"  ✓ 成功替换第{row_number}行的{replaced_count}个单元格")
            return replaced_count
            
        except Exception as e:
            logger.error(f"✗ 替换第{row_number}行数据时出错: {str(e)}", exc_info=True)
            return 0


# 文件未完成，继续在下一部分...


# ============================================================================
# ExcelSequenceProcessor主处理器 (Task 2)
# ============================================================================

class ExcelSequenceProcessor:
    """Excel序号处理器 - 协调整个处理流程"""
    
    def __init__(self, excel_path: str, ai_result: str, config: ProcessingConfig = None):
        """
        初始化处理器
        
        Args:
            excel_path: Excel文件路径
            ai_result: AI分析结果
            config: 处理配置
        """
        self.excel_path = excel_path
        self.ai_result = ai_result
        self.config = config or ProcessingConfig()
        self.workbook = None
        self.worksheet = None
        self.statistics = ProcessingStatistics()
    
    def process(self) -> ProcessingResult:
        """
        执行完整的处理流程
        
        Returns:
            ProcessingResult对象
        """
        start_time = time.time()
        warnings = []
        
        try:
            # 1. 提取所有表格
            logger.info("=" * 80)
            logger.info("开始处理Excel文件（基于序号列匹配）")
            logger.info(f"原文件: {self.excel_path}")
            logger.info("=" * 80)
            
            tables = TableExtractor.extract_all_tables(self.ai_result)
            
            if not tables:
                logger.error("✗ AI返回的内容中未找到Markdown表格")
                return ProcessingResult(
                    success=False,
                    error="AI返回的内容中未找到Markdown表格",
                    statistics=self.statistics
                )
            
            self.statistics.total_tables = len(tables)
            logger.info(f"准备处理 {len(tables)} 个表格")
            
            # 2. 加载Excel文件
            if not os.path.exists(self.excel_path):
                logger.error(f"✗ 原始文件不存在: {self.excel_path}")
                return ProcessingResult(
                    success=False,
                    error=f"原始文件不存在: {self.excel_path}",
                    statistics=self.statistics
                )
            
            try:
                logger.info("加载Excel文件...")
                self.workbook = load_workbook(self.excel_path)
                self.worksheet = self.workbook.active
                logger.info(f"✓ Excel文件加载成功 (工作表: {self.worksheet.title}, 行数: {self.worksheet.max_row})")
            except Exception as e:
                logger.error(f"✗ 加载Excel文件失败: {str(e)}")
                return ProcessingResult(
                    success=False,
                    error=f"加载Excel文件失败: {str(e)}",
                    statistics=self.statistics
                )
            
            # 3. 定位序号列
            logger.info("-" * 80)
            seq_col_info = SequenceColumnLocator.locate_sequence_column(
                self.worksheet,
                self.config.max_header_search_rows
            )
            
            if not seq_col_info:
                self.workbook.close()
                error_msg = f"在前{self.config.max_header_search_rows}行中未找到'序号'列，无法进行匹配"
                logger.error(f"✗ {error_msg}")
                return ProcessingResult(
                    success=False,
                    error=error_msg,
                    statistics=self.statistics
                )
            
            # 4. 构建序号匹配器
            logger.info("-" * 80)
            sequence_matcher = SequenceMatcher(
                self.worksheet,
                seq_col_info.column_index,
                seq_col_info.header_row
            )
            
            # 5. 处理每个表格
            # 按顺序处理所有包含序号列的表格，后面的表格会覆盖前面的相同序号行
            logger.info("-" * 80)
            
            # 统计包含序号列的表格数量
            tables_with_sequence_count = sum(1 for t in tables if t.has_sequence)
            
            if tables_with_sequence_count > 1:
                logger.info(f"检测到{tables_with_sequence_count}个包含序号列的表格")
                logger.info(f"将按顺序处理所有表格，后续表格会覆盖相同序号的行")
            
            # 按顺序处理所有表格
            for table_idx, table in enumerate(tables, 1):
                logger.info(f"[表格 {table_idx}/{len(tables)}] 开始处理...")
                
                try:
                    result = self.process_single_table(
                        table,
                        sequence_matcher,
                        seq_col_info.column_headers
                    )
                    
                    if result:
                        self.statistics.processed_tables += 1
                        logger.info(f"[表格 {table_idx}/{len(tables)}] ✓ 处理完成")
                    else:
                        self.statistics.skipped_tables += 1
                        logger.warning(f"[表格 {table_idx}/{len(tables)}] ⚠ 跳过该表格（不包含序号列）")
                        warnings.append(f"表格{table_idx}不包含序号列，已跳过")
                        
                except Exception as e:
                    logger.error(f"[表格 {table_idx}/{len(tables)}] ✗ 处理失败: {str(e)}")
                    self.statistics.skipped_tables += 1
                    warnings.append(f"表格{table_idx}处理失败: {str(e)}")
                
                logger.info("-" * 80)
            
            # 6. 保存文件
            if self.statistics.matched_rows == 0:
                self.workbook.close()
                logger.error("✗ 没有任何行被成功匹配和替换")
                return ProcessingResult(
                    success=False,
                    error="没有任何行被成功匹配和替换",
                    statistics=self.statistics,
                    warnings=warnings
                )
            
            try:
                logger.info("保存修改后的文件...")
                output_path = self._save_workbook()
                self.workbook.close()
                logger.info(f"✓ 文件保存成功: {output_path}")
            except Exception as e:
                self.workbook.close()
                logger.error(f"✗ 保存文件失败: {str(e)}")
                return ProcessingResult(
                    success=False,
                    error=f"保存文件失败: {str(e)}",
                    statistics=self.statistics,
                    warnings=warnings
                )
            
            # 7. 返回成功结果
            self.statistics.processing_time = time.time() - start_time
            
            # 输出最终统计信息
            logger.info("=" * 80)
            logger.info("处理完成 - 统计信息:")
            logger.info(f"  总表格数: {self.statistics.total_tables}")
            logger.info(f"  成功处理: {self.statistics.processed_tables} 个表格")
            logger.info(f"  跳过表格: {self.statistics.skipped_tables} 个表格")
            logger.info(f"  总数据行: {self.statistics.total_rows}")
            logger.info(f"  成功匹配: {self.statistics.matched_rows} 行 ({self.statistics.matched_rows/self.statistics.total_rows*100:.1f}%)" if self.statistics.total_rows > 0 else "  成功匹配: 0 行")
            logger.info(f"  跳过行数: {self.statistics.skipped_rows} 行")
            logger.info(f"  处理耗时: {self.statistics.processing_time:.2f} 秒")
            logger.info(f"  输出文件: {output_path}")
            logger.info("=" * 80)
            
            return ProcessingResult(
                success=True,
                output_path=output_path,
                filename=os.path.basename(output_path),
                statistics=self.statistics,
                warnings=warnings
            )
            
        except Exception as e:
            logger.error(f"✗ Excel处理失败: {str(e)}", exc_info=True)
            if self.workbook:
                self.workbook.close()
            
            return ProcessingResult(
                success=False,
                error=f"Excel处理失败: {str(e)}",
                statistics=self.statistics,
                warnings=warnings
            )
    
    def process_single_table(self,
                            table: TableData,
                            sequence_matcher: SequenceMatcher,
                            column_mapping: Dict[str, int]) -> bool:
        """
        处理单个表格
        
        Args:
            table: 表格数据
            sequence_matcher: 序号匹配器
            column_mapping: 列映射（列名到Excel列索引）
            
        Returns:
            是否处理成功
        """
        try:
            # 1. 检查表格是否包含序号列
            if not table.has_sequence:
                logger.warning("  ⚠ 表格不包含'序号'列，跳过")
                self.statistics.skipped_rows += len(table.rows)
                return False
            
            logger.info(f"  表格包含 {len(table.headers)} 列, {len(table.rows)} 行数据")
            logger.debug(f"  表头: {table.headers}")
            logger.info(f"  序号列位置: 第{table.sequence_col_index + 1}列")
            
            self.statistics.total_rows += len(table.rows)
            
            # 2. 处理每一行
            logger.info(f"  开始处理 {len(table.rows)} 行数据...")
            matched_in_table = 0
            skipped_in_table = 0
            
            for row_idx, row_data in enumerate(table.rows, 1):
                try:
                    # 提取序号值
                    sequence_value = row_data.get("序号", "")
                    
                    if not sequence_value or str(sequence_value).strip() == "":
                        logger.warning(f"  ⚠ 跳过第 {row_idx} 行: 序号为空")
                        self.statistics.skipped_rows += 1
                        skipped_in_table += 1
                        continue
                    
                    logger.debug(f"  处理第 {row_idx}/{len(table.rows)} 行 (序号: {sequence_value})...")
                    
                    # 查找匹配行
                    excel_row_num = sequence_matcher.find_row_by_sequence(sequence_value)
                    
                    if excel_row_num:
                        # 替换数据
                        replaced_count = DataReplacer.replace_row(
                            self.worksheet,
                            excel_row_num,
                            row_data,
                            column_mapping
                        )
                        
                        if replaced_count > 0:
                            logger.info(f"  ✓ 序号 {sequence_value} 匹配成功 -> Excel第{excel_row_num}行 (替换{replaced_count}列)")
                            self.statistics.matched_rows += 1
                            matched_in_table += 1
                        else:
                            logger.warning(f"  ⚠ 序号 {sequence_value} 找到但未替换任何列")
                            self.statistics.skipped_rows += 1
                            skipped_in_table += 1
                    else:
                        logger.warning(f"  ⚠ 跳过第 {row_idx} 行: 序号 {sequence_value} 在Excel中未找到")
                        self.statistics.skipped_rows += 1
                        skipped_in_table += 1
                        
                except Exception as e:
                    logger.error(f"  ✗ 处理第{row_idx}行时出错: {str(e)}")
                    self.statistics.skipped_rows += 1
                    skipped_in_table += 1
            
            # 输出该表格的处理统计
            logger.info(f"  表格处理统计: 成功匹配 {matched_in_table} 行, 跳过 {skipped_in_table} 行")
            
            return True
            
        except Exception as e:
            logger.error(f"  ✗ 表格处理失败: {str(e)}", exc_info=True)
            return False
    
    def _save_workbook(self) -> str:
        """
        保存工作簿
        
        Returns:
            保存后的文件路径
        """
        output_dir = 'uploads/modified'
        os.makedirs(output_dir, exist_ok=True)
        
        # 生成新文件名
        basename = os.path.basename(self.excel_path)
        name_without_ext, ext = os.path.splitext(basename)
        new_filename = f"{name_without_ext}(修改后).xlsx"
        
        # 构造完整输出路径
        output_path = os.path.join(output_dir, new_filename)
        
        # 处理文件名冲突
        counter = 1
        base_output_path = output_path
        while os.path.exists(output_path):
            name_without_ext, ext = os.path.splitext(base_output_path)
            output_path = f"{name_without_ext}_{counter}{ext}"
            counter += 1
        
        # 保存文件
        self.workbook.save(output_path)
        logger.info(f"文件已保存到: {output_path}")
        
        return output_path


# ============================================================================
# 主函数 (Task 3)
# ============================================================================

def modify_excel_by_sequence(original_path: str, ai_result: str, output_dir: str = 'uploads/modified',
                             config: ProcessingConfig = None) -> Dict:
    """
    主函数：执行基于序号列的Excel修改流程
    
    Args:
        original_path: 原Excel文件路径
        ai_result: AI返回的Markdown格式结果
        output_dir: 输出目录（暂未使用，保持向后兼容）
        config: 处理配置
        
    Returns:
        结果字典，包含success, output_path, filename, statistics或error信息
    """
    try:
        # 使用ExcelSequenceProcessor处理
        processor = ExcelSequenceProcessor(original_path, ai_result, config)
        result = processor.process()
        
        # 转换为字典格式（保持API兼容性）
        response = {
            'success': result.success,
            'statistics': {
                'total_tables': result.statistics.total_tables,
                'processed_tables': result.statistics.processed_tables,
                'skipped_tables': result.statistics.skipped_tables,
                'total_rows': result.statistics.total_rows,
                'matched_rows': result.statistics.matched_rows,
                'skipped_rows': result.statistics.skipped_rows,
                'processing_time': result.statistics.processing_time
            }
        }
        
        if result.success:
            response['output_path'] = result.output_path
            response['filename'] = result.filename
        else:
            response['error'] = result.error
        
        if result.warnings:
            response['warnings'] = result.warnings
        
        return response
        
    except Exception as e:
        logger.error(f"modify_excel_by_sequence函数出错: {str(e)}", exc_info=True)
        return {
            'success': False,
            'error': f"处理过程中发生错误: {str(e)}"
        }


def main():
    """命令行入口函数"""
    import io
    
    # 强制设置stdin/stdout为UTF-8编码
    sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding='utf-8')
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
    
    if len(sys.argv) < 2:
        result = {
            'success': False,
            'error': '参数不足。用法: python modify_excel_by_sequence.py <原文件路径> [输出目录] (AI结果从stdin读取)'
        }
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)
    
    original_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else 'uploads/modified'
    
    # 从stdin读取AI结果
    ai_result = sys.stdin.read()
    
    # 执行处理
    result = modify_excel_by_sequence(original_path, ai_result, output_dir)
    
    # 输出JSON结果
    print(json.dumps(result, ensure_ascii=False))
    sys.stdout.flush()
    
    # 根据结果设置退出码
    sys.exit(0 if result['success'] else 1)


if __name__ == '__main__':
    main()
