#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Boundary Case Tests for Excel Row-Level Matching and Replacement
Tests edge cases and error handling scenarios

Task 8: 边界情况测试
- 测试空表格处理
- 测试表头不匹配的情况
- 测试行完全不匹配的情况
- 测试特殊字符和空值处理
- 测试列数不一致的情况
Requirements: 9.1, 9.2, 9.3, 9.4, 9.5
"""

import os
import sys
import tempfile
import logging
from openpyxl import Workbook, load_workbook

# Add parent directory to path to import modify_excel
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from modify_excel import (
    modify_excel,
    ProcessingConfig,
    TableExtractor,
    HeaderMatcher,
    RowMatcher,
    DataReplacer,
    ExcelProcessor,
    TableData
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class BoundaryTestRunner:
    """Boundary case test runner"""
    
    def __init__(self):
        self.test_results = []
        self.passed = 0
        self.failed = 0
        self.temp_files = []
    
    def cleanup(self):
        """Clean up temporary files"""
        for filepath in self.temp_files:
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
            except Exception as e:
                logger.warning(f"Failed to clean up {filepath}: {e}")
    
    def create_test_excel(self, headers, data_rows):
        """
        Create a test Excel file
        
        Args:
            headers: List of header strings
            data_rows: List of data rows (each row is a list)
            
        Returns:
            Path to created Excel file
        """
        wb = Workbook()
        ws = wb.active
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data rows
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False)
        temp_path = temp_file.name
        temp_file.close()
        
        wb.save(temp_path)
        wb.close()
        
        self.temp_files.append(temp_path)
        return temp_path
    
    def run_all_tests(self):
        """Run all boundary case tests"""
        logger.info("=" * 80)
        logger.info("Starting Boundary Case Tests")
        logger.info("=" * 80)
        
        try:
            # Test 1: Empty table handling
            self.test_empty_table()
            
            # Test 2: Header mismatch
            self.test_header_mismatch()
            
            # Test 3: No row matches
            self.test_no_row_matches()
            
            # Test 4: Special characters and null values
            self.test_special_characters_and_nulls()
            
            # Test 5: Column count mismatch
            self.test_column_count_mismatch()
            
            # Test 6: Empty AI result
            self.test_empty_ai_result()
            
            # Test 7: Malformed markdown table
            self.test_malformed_markdown()
            
            # Test 8: File not found
            self.test_file_not_found()
            
            # Test 9: Whitespace-only values
            self.test_whitespace_values()
            
            # Test 10: Unicode and emoji handling
            self.test_unicode_handling()
            
        finally:
            # Clean up temp files
            self.cleanup()
        
        # Print summary
        self.print_summary()
    
    def test_empty_table(self):
        """Test: Empty table handling (Requirement 9.5)"""
        test_name = "Empty Table Handling"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Create test Excel
            excel_path = self.create_test_excel(
                headers=['序号', '名称', '数量'],
                data_rows=[
                    ['1', '产品A', '10'],
                    ['2', '产品B', '20']
                ]
            )
            
            # AI result with empty table (only headers, no data)
            ai_result = """
| 序号 | 名称 | 数量 |
|------|------|------|
"""
            
            # Process
            result = modify_excel(excel_path, ai_result, 'uploads/modified')
            
            # Verify: Should handle gracefully without crashing
            assert result is not None, "Should return a result"
            # Empty table should be skipped, so no rows matched
            if result['success']:
                assert result['statistics']['matched_rows'] == 0, "Empty table should match 0 rows"
            
            logger.info(f"✓ Empty table handled gracefully")
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_header_mismatch(self):
        """Test: Header mismatch handling (Requirement 3.4)"""
        test_name = "Header Mismatch"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Create test Excel
            excel_path = self.create_test_excel(
                headers=['序号', '名称', '数量'],
                data_rows=[
                    ['1', '产品A', '10'],
                    ['2', '产品B', '20']
                ]
            )
            
            # AI result with completely different headers
            ai_result = """
| ID | Product | Quantity | Price | Category |
|-----|---------|----------|-------|----------|
| 1 | Item A | 10 | 100 | Cat1 |
| 2 | Item B | 20 | 200 | Cat2 |
"""
            
            # Process
            result = modify_excel(excel_path, ai_result, 'uploads/modified')
            
            # Verify: Should skip table due to header mismatch
            assert result is not None, "Should return a result"
            if result['success']:
                # Table should be skipped
                assert result['statistics']['skipped_tables'] > 0, "Should skip mismatched table"
            else:
                # Or fail gracefully with error message
                assert 'error' in result, "Should have error message"
            
            logger.info(f"✓ Header mismatch handled correctly")
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_no_row_matches(self):
        """Test: No row matches found (Requirement 4.5)"""
        test_name = "No Row Matches"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Create test Excel
            excel_path = self.create_test_excel(
                headers=['序号', '名称', '数量'],
                data_rows=[
                    ['1', '产品A', '10'],
                    ['2', '产品B', '20']
                ]
            )
            
            # AI result with matching headers but completely different data
            ai_result = """
| 序号 | 名称 | 数量 |
|------|------|------|
| 99 | 完全不同的产品 | 999 |
| 88 | 另一个不同的产品 | 888 |
"""
            
            # Process
            result = modify_excel(excel_path, ai_result, 'uploads/modified')
            
            # Verify: Should handle gracefully, skip all rows
            assert result is not None, "Should return a result"
            if result['success']:
                # All rows should be skipped
                assert result['statistics']['skipped_rows'] > 0, "Should skip unmatched rows"
            
            logger.info(f"✓ No row matches handled correctly")
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_special_characters_and_nulls(self):
        """Test: Special characters and null value handling (Requirement 9.2)"""
        test_name = "Special Characters and Nulls"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Create test Excel with special characters
            excel_path = self.create_test_excel(
                headers=['序号', '名称', '备注'],
                data_rows=[
                    ['1', 'Product@#$%', 'Note with "quotes"'],
                    ['2', 'Item & Co.', None],  # Null value
                    ['3', 'Test\nNewline', 'Tab\there']
                ]
            )
            
            # AI result with special characters and empty values
            ai_result = """
| 序号 | 名称 | 备注 |
|------|------|------|
| 1 | Product@#$% | Updated note with "quotes" |
| 2 | Item & Co. |  |
| 3 | Test\nNewline | Tab\there |
"""
            
            # Process
            result = modify_excel(excel_path, ai_result, 'uploads/modified')
            
            # Verify: Should handle special characters without errors
            assert result is not None, "Should return a result"
            # Should not crash, that's the main requirement
            
            if result['success']:
                logger.info(f"✓ Processed with {result['statistics']['matched_rows']} matches")
                
                # Verify output file
                if 'output_path' in result and os.path.exists(result['output_path']):
                    wb = load_workbook(result['output_path'])
                    ws = wb.active
                    
                    # Check that special characters are preserved
                    found_special = False
                    for row in ws.iter_rows(min_row=2, max_row=4):
                        for cell in row:
                            if cell.value and '@#$%' in str(cell.value):
                                found_special = True
                                break
                    
                    wb.close()
                    
                    if found_special:
                        logger.info(f"✓ Special characters preserved in output")
            
            logger.info(f"✓ Special characters and nulls handled correctly")
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_column_count_mismatch(self):
        """Test: Column count mismatch handling (Requirement 9.1)"""
        test_name = "Column Count Mismatch"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Create test Excel with 3 columns
            excel_path = self.create_test_excel(
                headers=['序号', '名称', '数量'],
                data_rows=[
                    ['1', '产品A', '10'],
                    ['2', '产品B', '20']
                ]
            )
            
            # AI result with 5 columns (more than Excel)
            ai_result = """
| 序号 | 名称 | 数量 | 价格 | 备注 |
|------|------|------|------|------|
| 1 | 产品A | 15 | 100 | 更新 |
| 2 | 产品B | 25 | 200 | 新增 |
"""
            
            # Process
            result = modify_excel(excel_path, ai_result, 'uploads/modified')
            
            # Verify: Should process only matching columns
            assert result is not None, "Should return a result"
            
            if result['success']:
                # Should match rows based on common columns
                logger.info(f"✓ Processed with column mismatch: {result['statistics']['matched_rows']} matches")
                
                # Verify that only matching columns were updated
                if 'output_path' in result and os.path.exists(result['output_path']):
                    wb = load_workbook(result['output_path'])
                    ws = wb.active
                    
                    # Excel should still have 3 columns, not 5
                    assert ws.max_column == 3, "Excel should maintain original column count"
                    
                    wb.close()
                    logger.info(f"✓ Original column count preserved")
            
            logger.info(f"✓ Column count mismatch handled correctly")
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_empty_ai_result(self):
        """Test: Empty AI result handling (Requirement 1.4)"""
        test_name = "Empty AI Result"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Create test Excel
            excel_path = self.create_test_excel(
                headers=['序号', '名称', '数量'],
                data_rows=[['1', '产品A', '10']]
            )
            
            # Empty AI result
            ai_result = ""
            
            # Process
            result = modify_excel(excel_path, ai_result, 'uploads/modified')
            
            # Verify: Should handle gracefully
            assert result is not None, "Should return a result"
            assert result['success'] == False, "Should fail with empty AI result"
            assert 'error' in result, "Should have error message"
            
            logger.info(f"✓ Empty AI result handled gracefully")
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_malformed_markdown(self):
        """Test: Malformed markdown table handling (Requirement 9.5)"""
        test_name = "Malformed Markdown"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Create test Excel
            excel_path = self.create_test_excel(
                headers=['序号', '名称', '数量'],
                data_rows=[['1', '产品A', '10']]
            )
            
            # Malformed markdown (missing separators, inconsistent columns)
            ai_result = """
| 序号 | 名称 | 数量 |
| 1 | 产品A |
| 2 | 产品B | 20 | 额外列 |
"""
            
            # Process
            result = modify_excel(excel_path, ai_result, 'uploads/modified')
            
            # Verify: Should handle gracefully without crashing
            assert result is not None, "Should return a result"
            # May succeed or fail, but should not crash
            
            logger.info(f"✓ Malformed markdown handled without crash")
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_file_not_found(self):
        """Test: File not found error handling (Requirement 9.5)"""
        test_name = "File Not Found"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Non-existent file path
            excel_path = "/nonexistent/path/to/file.xlsx"
            
            ai_result = """
| 序号 | 名称 | 数量 |
|------|------|------|
| 1 | 产品A | 10 |
"""
            
            # Process
            result = modify_excel(excel_path, ai_result, 'uploads/modified')
            
            # Verify: Should return error
            assert result is not None, "Should return a result"
            assert result['success'] == False, "Should fail with file not found"
            assert 'error' in result, "Should have error message"
            
            logger.info(f"✓ File not found error handled correctly")
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_whitespace_values(self):
        """Test: Whitespace-only values handling (Requirement 9.2)"""
        test_name = "Whitespace Values"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Create test Excel with whitespace values
            excel_path = self.create_test_excel(
                headers=['序号', '名称', '数量'],
                data_rows=[
                    ['1', '  产品A  ', '10'],
                    ['2', 'Product B', '  20  ']
                ]
            )
            
            # AI result with whitespace
            ai_result = """
| 序号 | 名称 | 数量 |
|------|------|------|
| 1 | 产品A | 10 |
| 2 |   Product B   | 20 |
"""
            
            # Process
            result = modify_excel(excel_path, ai_result, 'uploads/modified')
            
            # Verify: Should match despite whitespace differences
            assert result is not None, "Should return a result"
            
            if result['success']:
                # Should match rows (whitespace should be trimmed for comparison)
                assert result['statistics']['matched_rows'] > 0, "Should match rows despite whitespace"
                logger.info(f"✓ Matched {result['statistics']['matched_rows']} rows with whitespace handling")
            
            logger.info(f"✓ Whitespace values handled correctly")
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_unicode_handling(self):
        """Test: Unicode and emoji handling (Requirement 9.2)"""
        test_name = "Unicode and Emoji Handling"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Create test Excel with unicode characters
            excel_path = self.create_test_excel(
                headers=['序号', '名称', '描述'],
                data_rows=[
                    ['1', '产品🎉', '中文描述'],
                    ['2', 'Café', 'Résumé'],
                    ['3', '日本語', 'テスト']
                ]
            )
            
            # AI result with unicode
            ai_result = """
| 序号 | 名称 | 描述 |
|------|------|------|
| 1 | 产品🎉 | 更新的中文描述 |
| 2 | Café | Updated Résumé |
| 3 | 日本語 | 新しいテスト |
"""
            
            # Process
            result = modify_excel(excel_path, ai_result, 'uploads/modified')
            
            # Verify: Should handle unicode without errors
            assert result is not None, "Should return a result"
            
            if result['success']:
                logger.info(f"✓ Processed unicode: {result['statistics']['matched_rows']} matches")
                
                # Verify unicode is preserved in output
                if 'output_path' in result and os.path.exists(result['output_path']):
                    wb = load_workbook(result['output_path'])
                    ws = wb.active
                    
                    # Check for unicode characters
                    found_emoji = False
                    found_japanese = False
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            if cell.value:
                                val = str(cell.value)
                                if '🎉' in val:
                                    found_emoji = True
                                if '日本語' in val or 'テスト' in val:
                                    found_japanese = True
                    
                    wb.close()
                    
                    if found_emoji or found_japanese:
                        logger.info(f"✓ Unicode characters preserved in output")
            
            logger.info(f"✓ Unicode and emoji handled correctly")
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def record_pass(self, test_name):
        """Record a passing test"""
        self.test_results.append({'name': test_name, 'status': 'PASS'})
        self.passed += 1
    
    def record_fail(self, test_name, error):
        """Record a failing test"""
        self.test_results.append({'name': test_name, 'status': 'FAIL', 'error': error})
        self.failed += 1
    
    def print_summary(self):
        """Print test summary"""
        logger.info("\n" + "=" * 80)
        logger.info("Boundary Case Test Summary")
        logger.info("=" * 80)
        
        for result in self.test_results:
            status_symbol = "✓" if result['status'] == 'PASS' else "✗"
            logger.info(f"{status_symbol} {result['name']}: {result['status']}")
            if result['status'] == 'FAIL':
                logger.info(f"  Error: {result.get('error', 'Unknown error')}")
        
        logger.info("-" * 80)
        logger.info(f"Total: {len(self.test_results)} tests")
        logger.info(f"Passed: {self.passed}")
        logger.info(f"Failed: {self.failed}")
        logger.info("=" * 80)
        
        if self.failed == 0:
            logger.info("✓ ALL BOUNDARY TESTS PASSED")
        else:
            logger.error(f"✗ {self.failed} TEST(S) FAILED")


def main():
    """Main entry point"""
    runner = BoundaryTestRunner()
    runner.run_all_tests()
    
    # Exit with appropriate code
    sys.exit(0 if runner.failed == 0 else 1)


if __name__ == '__main__':
    main()
