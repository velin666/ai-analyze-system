#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel智能行级匹配和替换脚本

实现基于AI分析结果的Excel文件智能修改功能：
- 表头匹配：通过50%阈值确定表格适用性
- 行级匹配：通过多列内容比对（至少2列）找到精确目标行
- 增量替换：仅替换匹配的行和列，保留其他数据
- 性能优化：行指针机制和回环搜索
- 错误恢复：完善的错误处理和恢复机制

版本: 2.0
"""

import sys
import json
import os
import time
import logging
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass, field
from openpyxl import load_workbook
from difflib import SequenceMatcher

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# 数据模型类 (Task 1.6)
# ============================================================================

@dataclass
class TableData:
    """表格数据结构"""
    headers: List[str]
    rows: List[List[str]]
    raw_text: str = ""


@dataclass
class HeaderMatchResult:
    """表头匹配结果"""
    matched: bool
    header_row: int = 0
    excel_headers: Dict[int, str] = field(default_factory=dict)
    column_mapping: Dict[int, int] = field(default_factory=dict)
    match_rate: float = 0.0


@dataclass
class RowMatchResult:
    """行匹配结果"""
    matched: bool
    row_number: int = 0
    matched_columns: int = 0
    matched_column_names: List[str] = field(default_factory=list)


@dataclass
class ProcessingStatistics:
    """处理统计信息"""
    total_tables: int = 0
    processed_tables: int = 0
    skipped_tables: int = 0
    total_rows: int = 0
    matched_rows: int = 0
    skipped_rows: int = 0
    processing_time: float = 0.0


@dataclass
class ProcessingResult:
    """处理结果"""
    success: bool
    output_path: Optional[str] = None
    filename: Optional[str] = None
    statistics: ProcessingStatistics = field(default_factory=ProcessingStatistics)
    error: Optional[str] = None
    warnings: List[str] = field(default_factory=list)


@dataclass
class ProcessingConfig:
    """处理配置"""
    row_match_threshold: int = 2
    header_match_threshold: float = 0.5
    enable_wraparound_search: bool = True
    max_search_distance: int = 1000
    preserve_formulas: bool = True
    log_level: str = "INFO"


@dataclass
class ErrorResponse:
    """错误响应"""
    code: str
    message: str
    user_message: str


# ============================================================================
# ErrorHandler类 (Task 4)
# ============================================================================

class ErrorHandler:
    """错误处理器 - 统一错误处理和恢复机制"""
    
    @staticmethod
    def handle_error(error: Exception, context: str) -> ErrorResponse:
        """
        统一错误处理
        
        Args:
            error: 异常对象
            context: 错误上下文描述
            
        Returns:
            ErrorResponse对象，包含错误代码、详细信息和用户友好提示
        """
        logger.error(f"{context}: {error}", exc_info=True)
        
        # 文件不存在错误
        if isinstance(error, FileNotFoundError):
            return ErrorResponse(
                code="FILE_NOT_FOUND",
                message=f"找不到指定的文件: {str(error)}",
                user_message="文件不存在，请检查文件路径是否正确"
            )
        
        # 权限错误
        elif isinstance(error, PermissionError):
            return ErrorResponse(
                code="PERMISSION_DENIED",
                message=f"没有文件访问权限: {str(error)}",
                user_message="无法访问文件，请检查文件权限或确保文件未被其他程序占用"
            )
        
        # 值错误（通常是数据格式问题）
        elif isinstance(error, ValueError):
            return ErrorResponse(
                code="INVALID_DATA",
                message=f"数据格式错误: {str(error)}",
                user_message="数据格式不正确，请检查输入数据"
            )
        
        # 类型错误
        elif isinstance(error, TypeError):
            return ErrorResponse(
                code="TYPE_ERROR",
                message=f"数据类型错误: {str(error)}",
                user_message="数据类型不匹配，请检查输入数据格式"
            )
        
        # IO错误（磁盘空间、文件锁等）
        elif isinstance(error, IOError) or isinstance(error, OSError):
            return ErrorResponse(
                code="IO_ERROR",
                message=f"文件操作错误: {str(error)}",
                user_message="文件读写失败，请检查磁盘空间或文件是否被占用"
            )
        
        # 编码错误
        elif isinstance(error, UnicodeDecodeError) or isinstance(error, UnicodeEncodeError):
            return ErrorResponse(
                code="ENCODING_ERROR",
                message=f"编码错误: {str(error)}",
                user_message="文件编码格式不支持，请确保文件使用UTF-8编码"
            )
        
        # 键错误（通常是数据结构问题）
        elif isinstance(error, KeyError):
            return ErrorResponse(
                code="KEY_ERROR",
                message=f"数据键不存在: {str(error)}",
                user_message="数据结构不完整，请检查输入数据"
            )
        
        # 索引错误
        elif isinstance(error, IndexError):
            return ErrorResponse(
                code="INDEX_ERROR",
                message=f"索引超出范围: {str(error)}",
                user_message="数据访问超出范围，请检查数据完整性"
            )
        
        # 未知错误
        else:
            return ErrorResponse(
                code="UNKNOWN_ERROR",
                message=f"未知错误: {str(error)}",
                user_message="处理过程中发生错误，请稍后重试或联系技术支持"
            )
    
    @staticmethod
    def handle_table_extraction_error(error: Exception) -> ErrorResponse:
        """
        处理表格提取错误
        
        Args:
            error: 异常对象
            
        Returns:
            ErrorResponse对象
        """
        return ErrorHandler.handle_error(error, "表格提取失败")
    
    @staticmethod
    def handle_table_parsing_error(error: Exception) -> ErrorResponse:
        """
        处理表格解析错误
        
        Args:
            error: 异常对象
            
        Returns:
            ErrorResponse对象
        """
        return ErrorHandler.handle_error(error, "表格解析失败")
    
    @staticmethod
    def handle_header_matching_error(error: Exception) -> ErrorResponse:
        """
        处理表头匹配错误
        
        Args:
            error: 异常对象
            
        Returns:
            ErrorResponse对象
        """
        return ErrorHandler.handle_error(error, "表头匹配失败")
    
    @staticmethod
    def handle_row_matching_error(error: Exception) -> ErrorResponse:
        """
        处理行匹配错误
        
        Args:
            error: 异常对象
            
        Returns:
            ErrorResponse对象
        """
        return ErrorHandler.handle_error(error, "行匹配失败")
    
    @staticmethod
    def handle_data_replacement_error(error: Exception) -> ErrorResponse:
        """
        处理数据替换错误
        
        Args:
            error: 异常对象
            
        Returns:
            ErrorResponse对象
        """
        return ErrorHandler.handle_error(error, "数据替换失败")
    
    @staticmethod
    def handle_file_operation_error(error: Exception) -> ErrorResponse:
        """
        处理文件操作错误
        
        Args:
            error: 异常对象
            
        Returns:
            ErrorResponse对象
        """
        return ErrorHandler.handle_error(error, "文件操作失败")
    
    @staticmethod
    def is_recoverable_error(error: Exception) -> bool:
        """
        判断错误是否可恢复（可以继续处理其他项）
        
        Args:
            error: 异常对象
            
        Returns:
            True表示可恢复，False表示不可恢复
        """
        # 以下错误类型被认为是可恢复的，可以继续处理其他项
        recoverable_errors = (
            ValueError,      # 数据格式错误
            TypeError,       # 类型错误
            KeyError,        # 键不存在
            IndexError,      # 索引错误
        )
        
        # 以下错误类型被认为是不可恢复的，应该停止处理
        non_recoverable_errors = (
            FileNotFoundError,   # 文件不存在
            PermissionError,     # 权限错误
            MemoryError,         # 内存错误
        )
        
        if isinstance(error, non_recoverable_errors):
            return False
        
        if isinstance(error, recoverable_errors):
            return True
        
        # 默认认为可恢复，尝试继续处理
        return True


# ============================================================================
# TableExtractor类 (Task 1.1)
# ============================================================================

class TableExtractor:
    """
    表格提取器 - 从Markdown文本中提取和解析表格
    
    负责从AI返回的Markdown格式文本中识别和提取所有表格，
    并将每个表格解析为结构化的TableData对象。
    
    表格识别规则：
    - 连续的包含'|'符号的行被识别为一个表格
    - 第一行为表头
    - 第二行为分隔符（跳过）
    - 第三行及以后为数据行
    """
    
    @staticmethod
    def extract_all_tables(markdown_text: str) -> List[str]:
        """
        提取AI返回内容中的所有Markdown表格
        
        识别规则：连续包含'|'符号的行被识别为一个表格
        
        Args:
            markdown_text: AI返回的Markdown格式文本
            
        Returns:
            表格文本列表，每个元素是一个完整的Markdown表格字符串
            
        Example:
            >>> text = "| 序号 | 名称 |\\n|---|---|\\n| 1 | 测试 |"
            >>> tables = TableExtractor.extract_all_tables(text)
            >>> len(tables)
            1
        """
        if not markdown_text or not isinstance(markdown_text, str):
            logger.warning("AI结果为空或格式不正确")
            return []
        
        try:
            logger.info("开始从AI结果中提取表格...")
            
            # 处理转义的换行符（某些API可能返回转义格式）
            if '\\n' in markdown_text:
                markdown_text = markdown_text.replace('\\n', '\n')
            
            tables = []
            lines = markdown_text.split('\n')
            current_table = []
            in_table = False
            
            for line in lines:
                stripped = line.strip()
                
                # 检查是否是表格行（包含|符号）
                if '|' in stripped:
                    if not in_table:
                        # 开始新表格
                        in_table = True
                        current_table = []
                    current_table.append(line)
                else:
                    # 遇到非表格行，如果之前在表格中则表格结束
                    if in_table and current_table:
                        tables.append('\n'.join(current_table))
                        current_table = []
                        in_table = False
            
            # 处理文件末尾的表格
            if in_table and current_table:
                tables.append('\n'.join(current_table))
            
            if tables:
                logger.info(f"✓ 成功提取 {len(tables)} 个表格")
            else:
                logger.warning("未在AI结果中找到任何Markdown表格")
            
            return tables
            
        except Exception as e:
            error_response = ErrorHandler.handle_table_extraction_error(e)
            logger.error(f"✗ {error_response.user_message}")
            return []
    
    @staticmethod
    def parse_table(table_text: str) -> Optional[TableData]:
        """
        解析单个Markdown表格为结构化数据
        
        Markdown表格格式：
        | 列1 | 列2 | 列3 |  <- 第1行：表头
        |-----|-----|-----|  <- 第2行：分隔符（跳过）
        | 值1 | 值2 | 值3 |  <- 第3行起：数据行
        
        Args:
            table_text: Markdown表格文本
            
        Returns:
            TableData对象，包含表头和数据行；如果解析失败返回None
        """
        if not table_text:
            return None
        
        try:
            lines = [line.strip() for line in table_text.split('\n') if line.strip()]
            
            if len(lines) < 2:
                return None
            
            # 解析表头（第一行）
            header_line = lines[0]
            headers = [cell.strip() for cell in header_line.split('|')]
            headers = [h for h in headers if h]  # 移除空字符串（首尾的|会产生空字符串）
            
            if not headers:
                return None
            
            # 解析数据行（从第三行开始，跳过分隔符行）
            rows = []
            for line in lines[2:]:
                cells = [cell.strip() for cell in line.split('|')]
                cells = [c for c in cells if c or c == '']
                
                # 移除首尾的空字符串（由首尾的|产生）
                if cells and cells[0] == '':
                    cells = cells[1:]
                if cells and cells[-1] == '':
                    cells = cells[:-1]
                
                # 过滤完全空的行
                if cells and any(cell for cell in cells):
                    # 确保列数与表头一致（补齐或截断）
                    while len(cells) < len(headers):
                        cells.append('')
                    rows.append(cells[:len(headers)])
            
            return TableData(headers=headers, rows=rows, raw_text=table_text)
            
        except Exception as e:
            error_response = ErrorHandler.handle_table_parsing_error(e)
            logger.error(f"✗ {error_response.user_message}")
            return None


# ============================================================================
# HeaderMatcher类 (Task 1.2)
# ============================================================================

class HeaderMatcher:
    """
    表头匹配器 - 匹配AI表格表头与Excel表头
    
    负责在Excel文件中查找与AI表格表头匹配的行，并创建列映射关系。
    
    匹配策略：
    1. 精确匹配：列名完全相同
    2. 包含匹配：一个列名包含另一个
    3. 模糊匹配：使用SequenceMatcher计算相似度（阈值0.6）
    
    匹配阈值：
    - 默认50%，即至少一半的AI表头列需要在Excel中找到匹配
    - 低于阈值的表格将被跳过
    """
    
    @staticmethod
    def match_header(ai_headers: List[str], 
                    excel_sheet,
                    match_threshold: float = 0.5) -> Optional[HeaderMatchResult]:
        """
        匹配表头并返回匹配结果
        
        Args:
            ai_headers: AI表格的表头列表
            excel_sheet: Excel工作表对象（openpyxl.worksheet.worksheet.Worksheet）
            match_threshold: 匹配阈值，范围[0.0, 1.0]，默认0.5表示50%
            
        Returns:
            HeaderMatchResult对象，包含匹配状态、列映射等信息；
            如果匹配失败返回matched=False的结果对象
        """
        try:
            logger.debug(f"开始匹配表头，AI表头: {ai_headers}")
            
            # 查找Excel中的表头行
            header_result = HeaderMatcher._find_header_row(excel_sheet)
            if not header_result:
                logger.warning("✗ 无法在Excel中找到表头行")
                return HeaderMatchResult(matched=False)
            
            header_row, excel_headers = header_result
            logger.debug(f"Excel表头行: 第{header_row}行, 列: {list(excel_headers.values())}")
            
            # 创建列映射
            column_mapping = HeaderMatcher.create_column_mapping(ai_headers, excel_headers)
            
            # 计算匹配率
            match_rate = len(column_mapping) / len(ai_headers) if ai_headers else 0
            
            # 判断是否达到匹配阈值
            if match_rate >= match_threshold:
                matched_pairs = [(ai_headers[ai_idx], excel_headers[excel_idx]) 
                                for ai_idx, excel_idx in column_mapping.items()]
                logger.info(f"✓ 表头匹配成功: {match_rate*100:.1f}% ({len(column_mapping)}/{len(ai_headers)}列)")
                logger.debug(f"  匹配的列对: {matched_pairs}")
                return HeaderMatchResult(
                    matched=True,
                    header_row=header_row,
                    excel_headers=excel_headers,
                    column_mapping=column_mapping,
                    match_rate=match_rate
                )
            else:
                logger.warning(f"✗ 表头匹配失败: 匹配率{match_rate*100:.1f}% < 阈值{match_threshold*100:.1f}%")
                logger.warning(f"  AI表头: {ai_headers}")
                logger.warning(f"  Excel表头: {list(excel_headers.values())}")
                logger.warning(f"  仅匹配了 {len(column_mapping)} 列，需要至少 {int(len(ai_headers) * match_threshold)} 列")
                return HeaderMatchResult(matched=False, match_rate=match_rate)
                
        except Exception as e:
            error_response = ErrorHandler.handle_header_matching_error(e)
            logger.error(f"✗ {error_response.user_message}")
            return HeaderMatchResult(matched=False)
    
    @staticmethod
    def create_column_mapping(ai_headers: List[str],
                            excel_headers: Dict[int, str]) -> Dict[int, int]:
        """
        创建列映射关系
        
        使用三级匹配策略：
        1. 精确匹配：列名完全相同（优先级最高）
        2. 包含匹配：一个列名包含另一个
        3. 模糊匹配：使用SequenceMatcher，相似度>0.6
        
        Args:
            ai_headers: AI表头列表（按顺序，0-based索引）
            excel_headers: Excel表头字典 {列索引(1-based): 列名}
            
        Returns:
            列映射字典 {AI列索引(0-based): Excel列索引(1-based)}
            
        Example:
            >>> ai_headers = ['序号', '名称', '数量']
            >>> excel_headers = {1: '序号', 2: '产品名称', 3: '数量'}
            >>> mapping = HeaderMatcher.create_column_mapping(ai_headers, excel_headers)
            >>> mapping
            {0: 1, 1: 2, 2: 3}
        """
        column_mapping = {}
        
        # 创建Excel列名到列索引的反向映射
        excel_name_to_idx = {name: idx for idx, name in excel_headers.items()}
        
        for ai_idx, ai_header in enumerate(ai_headers):
            ai_header_clean = ai_header.strip()
            
            # 1. 精确匹配
            if ai_header_clean in excel_name_to_idx:
                column_mapping[ai_idx] = excel_name_to_idx[ai_header_clean]
                continue
            
            # 2. 包含匹配
            for excel_name, excel_idx in excel_name_to_idx.items():
                if ai_header_clean in excel_name or excel_name in ai_header_clean:
                    column_mapping[ai_idx] = excel_idx
                    break
            
            if ai_idx in column_mapping:
                continue
            
            # 3. 模糊匹配
            best_match = None
            best_ratio = 0.6
            
            for excel_name, excel_idx in excel_name_to_idx.items():
                ratio = SequenceMatcher(None, ai_header_clean, excel_name).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_match = excel_idx
            
            if best_match is not None:
                column_mapping[ai_idx] = best_match
        
        return column_mapping
    
    @staticmethod
    def _find_header_row(worksheet) -> Optional[Tuple[int, Dict[int, str]]]:
        """
        定位Excel表头行
        
        使用启发式算法：
        1. 扫描前20行
        2. 计算每行的得分 = 关键词数量 * 10 + 非空单元格数
        3. 返回得分最高的行
        
        Args:
            worksheet: openpyxl的worksheet对象
            
        Returns:
            (表头行号, 表头字典{列索引: 列名})，如果找不到返回None
        """
        max_scan_rows = min(20, worksheet.max_row)
        # 常见的表头关键词
        header_keywords = ['序号', '名称', '品牌', '型号', '数量', '单位', '备注', 
                          'ERP', '识别码', '编号', '规格', '尺寸', '价格', '金额']
        
        best_match = None
        best_score = 0
        
        for row_num in range(1, max_scan_rows + 1):
            row = worksheet[row_num]
            
            # 收集非空单元格
            non_empty_cells = []
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is not None and str(cell.value).strip():
                    non_empty_cells.append((col_idx, str(cell.value).strip()))
            
            # 至少需要3个非空单元格才可能是表头
            if len(non_empty_cells) < 3:
                continue
            
            # 计算包含关键词的单元格数量
            keyword_count = 0
            for _, cell_value in non_empty_cells:
                for keyword in header_keywords:
                    if keyword in cell_value:
                        keyword_count += 1
                        break
            
            # 计算得分：关键词权重更高
            score = keyword_count * 10 + len(non_empty_cells)
            
            if score > best_score:
                best_score = score
                best_match = (row_num, {col_idx: header for col_idx, header in non_empty_cells})
            
            # 如果找到明显的表头行（3个以上关键词，5个以上列），立即返回
            if keyword_count >= 3 and len(non_empty_cells) >= 5:
                return best_match
        
        return best_match


# ============================================================================
# RowMatcher类 (Task 1.3)
# ============================================================================

class RowMatcher:
    """
    行匹配器 - 通过多列内容比对找到匹配的行
    
    核心功能：
    - 在Excel文件中查找与AI数据行匹配的目标行
    - 使用行指针优化查找性能（避免重复扫描）
    - 支持回环搜索提高匹配率
    
    匹配规则：
    - 至少N列（默认2列）内容完全一致才认为匹配
    - 比较时忽略大小写和前后空白
    
    性能优化：
    - 行指针机制：记录上次匹配位置，下次从该位置开始搜索
    - 回环搜索：第一次搜索失败后，从表头重新搜索到指针位置
    - 早期退出：找到匹配立即返回
    """
    
    def __init__(self, match_threshold: int = 2):
        """
        初始化行匹配器
        
        Args:
            match_threshold: 匹配阈值（至少需要匹配的列数），默认2
        """
        self.match_threshold = match_threshold
        self.current_pointer = 0  # 当前搜索指针位置
    
    def find_matching_row(self,
                         ai_row: List[str],
                         excel_sheet,
                         column_mapping: Dict[int, int],
                         start_row: int,
                         header_row: int,
                         enable_wraparound: bool = True) -> Optional[RowMatchResult]:
        """
        找到匹配的行号
        
        搜索策略：
        1. 第一次搜索：从当前指针位置到文件末尾
        2. 回环搜索（可选）：从表头下一行到当前指针位置
        3. 找到匹配后更新指针位置
        
        Args:
            ai_row: AI数据行（列表）
            excel_sheet: Excel工作表对象
            column_mapping: 列映射字典 {AI列索引: Excel列索引}
            start_row: 数据开始行号（通常是表头行+1）
            header_row: 表头行号
            enable_wraparound: 是否启用回环搜索，默认True
            
        Returns:
            RowMatchResult对象，包含匹配状态、行号、匹配列数等信息；
            如果未找到返回None
        """
        try:
            search_start = self.current_pointer if self.current_pointer > start_row else start_row
            logger.debug(f"    搜索范围: 第{search_start}行 到 第{excel_sheet.max_row}行")
            
            # 第一次搜索：从当前指针到文件末尾
            result = self._search_range(
                ai_row, excel_sheet, column_mapping,
                search_start,
                excel_sheet.max_row
            )
            
            if result:
                self.current_pointer = result.row_number
                return result
            
            # 回环搜索：从表头下一行到当前指针
            if enable_wraparound and self.current_pointer > start_row:
                logger.debug(f"    第一次搜索未找到，执行回环搜索: 第{start_row}行 到 第{self.current_pointer - 1}行")
                result = self._search_range(
                    ai_row, excel_sheet, column_mapping,
                    start_row, self.current_pointer - 1
                )
                
                if result:
                    self.current_pointer = result.row_number
                    logger.debug(f"    ✓ 回环搜索成功找到匹配")
                    return result
                else:
                    logger.debug(f"    ✗ 回环搜索也未找到匹配")
            
            return None
            
        except Exception as e:
            error_response = ErrorHandler.handle_row_matching_error(e)
            logger.error(f"✗ {error_response.user_message}")
            return None
    
    def _search_range(self,
                     ai_row: List[str],
                     excel_sheet,
                     column_mapping: Dict[int, int],
                     start_row: int,
                     end_row: int) -> Optional[RowMatchResult]:
        """
        在指定范围内搜索匹配行
        
        Args:
            ai_row: AI数据行
            excel_sheet: Excel工作表对象
            column_mapping: 列映射字典
            start_row: 开始行号
            end_row: 结束行号
            
        Returns:
            RowMatchResult对象，如果未找到返回None
        """
        for row_num in range(start_row, end_row + 1):
            excel_row = excel_sheet[row_num]
            matched_count, matched_names = self.compare_rows(
                ai_row, excel_row, column_mapping
            )
            
            if matched_count >= self.match_threshold:
                logger.info(f"  ✓ 行匹配成功: Excel第{row_num}行 (匹配{matched_count}列: {', '.join(matched_names)})")
                logger.debug(f"    AI数据: {ai_row[:3]}..." if len(ai_row) > 3 else f"    AI数据: {ai_row}")
                return RowMatchResult(
                    matched=True,
                    row_number=row_num,
                    matched_columns=matched_count,
                    matched_column_names=matched_names
                )
        
        return None
    
    @staticmethod
    def compare_rows(ai_row: List[str],
                    excel_row,
                    column_mapping: Dict[int, int]) -> Tuple[int, List[str]]:
        """
        比较两行数据，返回匹配的列数和列名
        
        比较规则：
        - 将两个值转换为字符串
        - 去除前后空白
        - 忽略大小写进行比较
        - 完全相同才算匹配
        
        Args:
            ai_row: AI数据行（列表）
            excel_row: Excel行对象（openpyxl的Row对象）
            column_mapping: 列映射字典 {AI列索引: Excel列索引}
            
        Returns:
            元组 (匹配的列数, 匹配的列名列表)
            
        Example:
            >>> ai_row = ['1', '测试', '10']
            >>> # excel_row 包含相同数据
            >>> count, names = RowMatcher.compare_rows(ai_row, excel_row, {0:1, 1:2, 2:3})
            >>> count
            3
        """
        matched_count = 0
        matched_names = []
        
        for ai_col_idx, excel_col_idx in column_mapping.items():
            if ai_col_idx >= len(ai_row):
                continue
            
            ai_value = str(ai_row[ai_col_idx]).strip()
            excel_cell = excel_row[excel_col_idx - 1]  # openpyxl行是0-based索引
            excel_value = str(excel_cell.value).strip() if excel_cell.value is not None else ""
            
            # 比较值（忽略大小写和空白）
            if ai_value.lower() == excel_value.lower():
                matched_count += 1
                matched_names.append(f"列{excel_col_idx}")
        
        return matched_count, matched_names


# ============================================================================
# DataReplacer类 (Task 1.4)
# ============================================================================

class DataReplacer:
    """
    数据替换器 - 替换Excel中的行数据
    
    负责将AI数据行的内容精确替换到Excel文件的目标行。
    
    替换策略：
    - 仅替换列映射中指定的列
    - 未映射的列保持原样
    - 保留原有单元格格式
    - 处理空值和特殊字符
    """
    
    @staticmethod
    def replace_row(excel_sheet,
                   row_number: int,
                   ai_row: List[str],
                   column_mapping: Dict[int, int]) -> None:
        """
        替换指定行的数据
        
        替换过程：
        1. 根据列映射确定要替换的列
        2. 将AI数据写入对应的Excel单元格
        3. 保留未映射列的原始数据
        4. 处理空值（转换为空字符串）
        
        Args:
            excel_sheet: Excel工作表对象
            row_number: 要替换的行号（1-based）
            ai_row: AI数据行（列表）
            column_mapping: 列映射字典 {AI列索引: Excel列索引}
        """
        try:
            excel_row = excel_sheet[row_number]
            replaced_count = 0
            
            # 仅替换映射的列
            for ai_col_idx, excel_col_idx in column_mapping.items():
                if ai_col_idx < len(ai_row):
                    ai_value = ai_row[ai_col_idx]
                    
                    # 处理空值和特殊字符
                    if ai_value is None or ai_value == '':
                        ai_value = ''
                    else:
                        ai_value = str(ai_value).strip()
                    
                    # 写入数据（保留原有格式）
                    cell = excel_row[excel_col_idx - 1]  # openpyxl行是0-based索引
                    old_value = cell.value
                    cell.value = ai_value
                    replaced_count += 1
                    
                    if old_value != ai_value:
                        logger.debug(f"    列{excel_col_idx}: '{old_value}' -> '{ai_value}'")
            
            logger.debug(f"  ✓ 成功替换第{row_number}行的{replaced_count}个单元格")
            
        except Exception as e:
            error_response = ErrorHandler.handle_data_replacement_error(e)
            logger.error(f"✗ 替换第{row_number}行数据时出错: {error_response.user_message}")
            # 对于数据替换错误，如果可恢复则不抛出异常，让调用者决定如何处理
            if not ErrorHandler.is_recoverable_error(e):
                raise


# ============================================================================
# ExcelProcessor主处理器 (Task 1.5)
# ============================================================================

class ExcelProcessor:
    """Excel处理器 - 协调整个处理流程"""
    
    def __init__(self, excel_path: str, ai_result: str, config: ProcessingConfig = None):
        """
        初始化处理器
        
        Args:
            excel_path: Excel文件路径
            ai_result: AI分析结果
            config: 处理配置
        """
        self.excel_path = excel_path
        self.ai_result = ai_result
        self.config = config or ProcessingConfig()
        self.workbook = None
        self.worksheet = None
        self.statistics = ProcessingStatistics()
    
    def process(self) -> ProcessingResult:
        """
        执行完整的处理流程
        
        Returns:
            ProcessingResult对象
        """
        start_time = time.time()
        warnings = []
        
        try:
            # 1. 提取所有表格
            logger.info("=" * 80)
            logger.info("开始处理Excel文件")
            logger.info(f"原文件: {self.excel_path}")
            logger.info("=" * 80)
            
            tables_text = TableExtractor.extract_all_tables(self.ai_result)
            
            if not tables_text:
                logger.error("✗ AI返回的内容中未找到Markdown表格")
                return ProcessingResult(
                    success=False,
                    error="AI返回的内容中未找到Markdown表格",
                    statistics=self.statistics
                )
            
            self.statistics.total_tables = len(tables_text)
            logger.info(f"准备处理 {len(tables_text)} 个表格")
            
            # 2. 加载Excel文件
            if not os.path.exists(self.excel_path): not os.path.exists(self.excel_path):
                logger.error(f"✗ 原始文件不存在: {self.excel_path}")
                return ProcessingResult(
                    success=False,
                    error=f"原始文件不存在: {self.excel_path}",
                    statistics=self.statistics
                )
            
            try:
                logger.info("加载Excel文件...")
                self.workbook = load_workbook(self.excel_path)
                self.worksheet = self.workbook.active
                logger.info(f"✓ Excel文件加载成功 (工作表: {self.worksheet.title}, 行数: {self.worksheet.max_row})")
            except Exception as e:
                error_response = ErrorHandler.handle_file_operation_error(e)
                logger.error(f"✗ {error_response.user_message}")
                return ProcessingResult(
                    success=False,
                    error=error_response.user_message,
                    statistics=self.statistics
                )
            
            # 3. 处理每个表格
            logger.info("-" * 80)
            for table_idx, table_text in enumerate(tables_text, 1):
                logger.info(f"[表格 {table_idx}/{len(tables_text)}] 开始处理...")
                
                try:
                    result = self.process_single_table(table_text)
                    
                    if result:
                        self.statistics.processed_tables += 1
                        logger.info(f"[表格 {table_idx}/{len(tables_text)}] ✓ 处理完成")
                    else:
                        self.statistics.skipped_tables += 1
                        logger.warning(f"[表格 {table_idx}/{len(tables_text)}] ⚠ 跳过该表格（表头匹配失败或无数据）")
                        warnings.append(f"表格{table_idx}处理失败")
                        
                except Exception as e:
                    # 使用错误处理器处理异常
                    error_response = ErrorHandler.handle_error(e, f"表格{table_idx}处理")
                    logger.error(f"[表格 {table_idx}/{len(tables_text)}] ✗ {error_response.user_message}")
                    self.statistics.skipped_tables += 1
                    warnings.append(f"表格{table_idx}: {error_response.user_message}")
                    
                    # 如果是不可恢复的错误，停止处理
                    if not ErrorHandler.is_recoverable_error(e):
                        logger.error("遇到不可恢复的错误，停止处理")
                        if self.workbook:
                            self.workbook.close()
                        return ProcessingResult(
                            success=False,
                            error=error_response.user_message,
                            statistics=self.statistics,
                            warnings=warnings
                        )
                
                logger.info("-" * 80)
            
            # 4. 保存文件
            if self.statistics.matched_rows == 0:
                self.workbook.close()
                logger.error("✗ 没有任何行被成功匹配和替换")
                return ProcessingResult(
                    success=False,
                    error="没有任何行被成功匹配和替换",
                    statistics=self.statistics,
                    warnings=warnings
                )
            
            try:
                logger.info("保存修改后的文件...")
                output_path = self._save_workbook()
                self.workbook.close()
                logger.info(f"✓ 文件保存成功: {output_path}")
            except Exception as e:
                self.workbook.close()
                error_response = ErrorHandler.handle_file_operation_error(e)
                logger.error(f"✗ {error_response.user_message}")
                return ProcessingResult(
                    success=False,
                    error=error_response.user_message,
                    statistics=self.statistics,
                    warnings=warnings
                )
            
            # 5. 返回成功结果
            self.statistics.processing_time = time.time() - start_time
            
            # 输出最终统计信息
            logger.info("=" * 80)
            logger.info("处理完成 - 统计信息:")
            logger.info(f"  总表格数: {self.statistics.total_tables}")
            logger.info(f"  成功处理: {self.statistics.processed_tables} 个表格")
            logger.info(f"  跳过表格: {self.statistics.skipped_tables} 个表格")
            logger.info(f"  总数据行: {self.statistics.total_rows}")
            logger.info(f"  成功匹配: {self.statistics.matched_rows} 行 ({self.statistics.matched_rows/self.statistics.total_rows*100:.1f}%)" if self.statistics.total_rows > 0 else "  成功匹配: 0 行")
            logger.info(f"  跳过行数: {self.statistics.skipped_rows} 行")
            logger.info(f"  处理耗时: {self.statistics.processing_time:.2f} 秒")
            logger.info(f"  输出文件: {output_path}")
            logger.info("=" * 80)
            
            return ProcessingResult(
                success=True,
                output_path=output_path,
                filename=os.path.basename(output_path),
                statistics=self.statistics,
                warnings=warnings
            )
            
        except Exception as e:
            error_response = ErrorHandler.handle_error(e, "Excel处理")
            logger.error(f"✗ {error_response.user_message}")
            if self.workbook:
                self.workbook.close()
            
            return ProcessingResult(
                success=False,
                error=error_response.user_message,
                statistics=self.statistics,
                warnings=warnings
            )
    
    def process_single_table(self, table_text: str) -> bool:
        """
        处理单个表格
        
        Args:
            table_text: 表格文本
            
        Returns:
            是否处理成功
        """
        try:
            # 1. 解析表格
            logger.debug("  解析表格数据...")
            table_data = TableExtractor.parse_table(table_text)
            if not table_data:
                logger.warning("  ✗ 表格解析失败")
                return False
            
            logger.info(f"  表格包含 {len(table_data.headers)} 列, {len(table_data.rows)} 行数据")
            logger.debug(f"  表头: {table_data.headers}")
            self.statistics.total_rows += len(table_data.rows)
            
            # 2. 匹配表头
            logger.info("  开始匹配表头...")
            header_result = HeaderMatcher.match_header(
                table_data.headers,
                self.worksheet,
                self.config.header_match_threshold
            )
            
            if not header_result or not header_result.matched:
                logger.warning(f"  ⚠ 跳过该表格: 表头匹配失败")
                logger.warning(f"    AI表头: {table_data.headers}")
                self.statistics.skipped_rows += len(table_data.rows)
                return False
            
            # 3. 初始化行匹配器
            logger.info(f"  开始处理 {len(table_data.rows)} 行数据...")
            row_matcher = RowMatcher(self.config.row_match_threshold)
            row_matcher.current_pointer = header_result.header_row + 1
            
            matched_in_table = 0
            skipped_in_table = 0
            
            # 4. 处理每一行
            for row_idx, ai_row in enumerate(table_data.rows, 1):
                try:
                    logger.debug(f"  处理第 {row_idx}/{len(table_data.rows)} 行...")
                    
                    # 查找匹配行
                    match_result = row_matcher.find_matching_row(
                        ai_row,
                        self.worksheet,
                        header_result.column_mapping,
                        header_result.header_row + 1,
                        header_result.header_row,
                        self.config.enable_wraparound_search
                    )
                    
                    if match_result and match_result.matched:
                        # 替换数据
                        DataReplacer.replace_row(
                            self.worksheet,
                            match_result.row_number,
                            ai_row,
                            header_result.column_mapping
                        )
                        self.statistics.matched_rows += 1
                        matched_in_table += 1
                    else:
                        logger.warning(f"  ⚠ 跳过第 {row_idx} 行: 未找到匹配的Excel行")
                        logger.debug(f"    AI数据: {ai_row}")
                        self.statistics.skipped_rows += 1
                        skipped_in_table += 1
                        
                except Exception as e:
                    error_response = ErrorHandler.handle_error(e, f"第{row_idx}行处理")
                    logger.error(f"  ✗ {error_response.user_message}")
                    self.statistics.skipped_rows += 1
                    skipped_in_table += 1
            
            # 输出该表格的处理统计
            logger.info(f"  表格处理统计: 成功匹配 {matched_in_table} 行, 跳过 {skipped_in_table} 行")
            
            return True
            
        except Exception as e:
            error_response = ErrorHandler.handle_error(e, "表格处理")
            logger.error(f"  ✗ {error_response.user_message}")
            return False
    
    def _save_workbook(self) -> str:
        """
        保存工作簿
        
        Returns:
            保存后的文件路径
        """
        output_dir = 'uploads/modified'
        os.makedirs(output_dir, exist_ok=True)
        
        # 生成新文件名
        basename = os.path.basename(self.excel_path)
        name_without_ext, ext = os.path.splitext(basename)
        new_filename = f"{name_without_ext}(修改后).xlsx"
        
        # 构造完整输出路径
        output_path = os.path.join(output_dir, new_filename)
        
        # 处理文件名冲突
        counter = 1
        base_output_path = output_path
        while os.path.exists(output_path):
            name_without_ext, ext = os.path.splitext(base_output_path)
            output_path = f"{name_without_ext}_{counter}{ext}"
            counter += 1
        
        # 保存文件
        self.workbook.save(output_path)
        logger.info(f"文件已保存到: {output_path}")
        
        return output_path


# ============================================================================
# 主函数
# ============================================================================

def modify_excel(original_path: str, ai_result: str, output_dir: str = 'uploads/modified', 
                config: ProcessingConfig = None) -> Dict:
    """
    主函数：执行完整的Excel修改流程（使用新的行级匹配逻辑）
    
    Args:
        original_path: 原Excel文件路径
        ai_result: AI返回的Markdown格式结果
        output_dir: 输出目录（暂未使用，保持向后兼容）
        config: 处理配置
        
    Returns:
        结果字典，包含success, output_path, filename, statistics或error信息
    """
    try:
        # 使用新的ExcelProcessor处理
        processor = ExcelProcessor(original_path, ai_result, config)
        result = processor.process()
        
        # 转换为字典格式（保持API兼容性）
        response = {
            'success': result.success,
            'statistics': {
                'total_tables': result.statistics.total_tables,
                'processed_tables': result.statistics.processed_tables,
                'skipped_tables': result.statistics.skipped_tables,
                'total_rows': result.statistics.total_rows,
                'matched_rows': result.statistics.matched_rows,
                'skipped_rows': result.statistics.skipped_rows,
                'processing_time': result.statistics.processing_time
            }
        }
        
        if result.success:
            response['output_path'] = result.output_path
            response['filename'] = result.filename
        else:
            response['error'] = result.error
        
        if result.warnings:
            response['warnings'] = result.warnings
        
        return response
        
    except Exception as e:
        error_response = ErrorHandler.handle_error(e, "modify_excel函数")
        logger.error(f"modify_excel函数出错: {error_response.user_message}")
        return {
            'success': False,
            'error': error_response.user_message
        }


def main():
    """命令行入口函数"""
    import io
    
    # 强制设置stdin/stdout为UTF-8编码
    sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding='utf-8')
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
    
    if len(sys.argv) < 2:
        result = {
            'success': False,
            'error': '参数不足。用法: python modify_excel.py <原文件路径> [输出目录] (AI结果从stdin读取)'
        }
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)
    
    original_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else 'uploads/modified'
    
    # 从stdin读取AI结果
    ai_result = sys.stdin.read()
    
    # 执行处理
    result = modify_excel(original_path, ai_result, output_dir)
    
    # 输出JSON结果
    print(json.dumps(result, ensure_ascii=False))
    sys.stdout.flush()
    
    # 根据结果设置退出码
    sys.exit(0 if result['success'] else 1)


if __name__ == '__main__':
    main()
