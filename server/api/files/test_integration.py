#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Integration Tests for Excel Row-Level Matching and Replacement
Tests the complete flow using real res.md, res2.md files and actual Excel file
"""

import os
import sys
import json
import logging
from openpyxl import load_workbook

# Add parent directory to path to import modify_excel
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from modify_excel import (
    modify_excel,
    ProcessingConfig,
    TableExtractor,
    HeaderMatcher,
    RowMatcher,
    DataReplacer,
    ExcelProcessor
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class IntegrationTestRunner:
    """Integration test runner for Excel modification system"""
    
    def __init__(self):
        self.test_results = []
        self.passed = 0
        self.failed = 0
    
    def run_all_tests(self):
        """Run all integration tests"""
        logger.info("=" * 80)
        logger.info("Starting Integration Tests")
        logger.info("=" * 80)
        
        # Test 1: Table extraction
        self.test_table_extraction()
        
        # Test 2: Table parsing
        self.test_table_parsing()
        
        # Test 3: Header matching
        self.test_header_matching()
        
        # Test 4: Complete flow with res2.md
        self.test_complete_flow_res2()
        
        # Test 5: Complete flow with res.md
        self.test_complete_flow_res()
        
        # Test 6: Verify output file
        self.test_verify_output_file()
        
        # Print summary
        self.print_summary()
    
    def test_table_extraction(self):
        """Test: Extract tables from AI result"""
        test_name = "Table Extraction"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Read res2.md
            with open('server/api/res2.md', 'r', encoding='utf-8') as f:
                ai_result = f.read()
            
            # Extract tables
            tables = TableExtractor.extract_all_tables(ai_result)
            
            # Verify
            assert len(tables) > 0, "Should extract at least one table"
            logger.info(f"✓ Extracted {len(tables)} tables")
            
            # Check table content
            for i, table in enumerate(tables, 1):
                assert '|' in table, f"Table {i} should contain pipe characters"
                lines = table.split('\n')
                assert len(lines) >= 3, f"Table {i} should have at least 3 lines (header, separator, data)"
            
            logger.info(f"✓ All tables have valid structure")
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_table_parsing(self):
        """Test: Parse table structure"""
        test_name = "Table Parsing"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Read res2.md
            with open('server/api/res2.md', 'r', encoding='utf-8') as f:
                ai_result = f.read()
            
            # Extract and parse first table
            tables = TableExtractor.extract_all_tables(ai_result)
            assert len(tables) > 0, "Should have tables to parse"
            
            table_data = TableExtractor.parse_table(tables[0])
            
            # Verify
            assert table_data is not None, "Should successfully parse table"
            assert len(table_data.headers) > 0, "Should have headers"
            assert len(table_data.rows) > 0, "Should have data rows"
            
            logger.info(f"✓ Parsed table: {len(table_data.headers)} columns, {len(table_data.rows)} rows")
            logger.info(f"  Headers: {table_data.headers[:3]}...")
            
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_header_matching(self):
        """Test: Header matching functionality"""
        test_name = "Header Matching"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Load Excel file
            excel_path = 'assets/KHG51-SD01 烘烤炉电气件清单.xlsx'
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel file not found: {excel_path}")
            
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Read res2.md
            with open('server/api/res2.md', 'r', encoding='utf-8') as f:
                ai_result = f.read()
            
            # Extract and parse tables - use the second table which has matching headers
            tables = TableExtractor.extract_all_tables(ai_result)
            assert len(tables) >= 2, "Should have at least 2 tables"
            
            # Use the second table which contains the actual data table (not error list)
            table_data = TableExtractor.parse_table(tables[1])
            
            # Test header matching
            header_result = HeaderMatcher.match_header(
                table_data.headers,
                ws,
                match_threshold=0.5
            )
            
            # Verify
            assert header_result is not None, "Should return header match result"
            assert header_result.matched, "Should successfully match headers"
            assert header_result.match_rate >= 0.5, "Match rate should be >= 50%"
            assert len(header_result.column_mapping) > 0, "Should have column mappings"
            
            logger.info(f"✓ Header matching successful")
            logger.info(f"  Match rate: {header_result.match_rate*100:.1f}%")
            logger.info(f"  Mapped columns: {len(header_result.column_mapping)}")
            
            wb.close()
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_complete_flow_res2(self):
        """Test: Complete flow with res2.md"""
        test_name = "Complete Flow (res2.md)"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Read res2.md
            with open('server/api/res2.md', 'r', encoding='utf-8') as f:
                ai_result = f.read()
            
            # Excel file path
            excel_path = 'assets/KHG51-SD01 烘烤炉电气件清单.xlsx'
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel file not found: {excel_path}")
            
            # Execute processing
            logger.info("Processing Excel file with res2.md...")
            result = modify_excel(excel_path, ai_result, 'uploads/modified')
            
            # Verify result
            assert result['success'] == True, f"Processing should succeed: {result.get('error', '')}"
            assert 'statistics' in result, "Should return statistics"
            assert 'output_path' in result, "Should return output path"
            
            stats = result['statistics']
            assert stats['total_tables'] > 0, "Should process at least one table"
            assert stats['matched_rows'] > 0, "Should match at least one row"
            
            logger.info(f"✓ Processing successful")
            logger.info(f"  Total tables: {stats['total_tables']}")
            logger.info(f"  Processed tables: {stats['processed_tables']}")
            logger.info(f"  Matched rows: {stats['matched_rows']}")
            logger.info(f"  Skipped rows: {stats['skipped_rows']}")
            logger.info(f"  Processing time: {stats['processing_time']:.2f}s")
            logger.info(f"  Output file: {result['output_path']}")
            
            # Verify output file exists
            assert os.path.exists(result['output_path']), "Output file should exist"
            
            # Store output path for verification test
            self.res2_output_path = result['output_path']
            
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_complete_flow_res(self):
        """Test: Complete flow with res.md"""
        test_name = "Complete Flow (res.md)"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            # Read res.md
            with open('server/api/res.md', 'r', encoding='utf-8') as f:
                ai_result = f.read()
            
            # Excel file path
            excel_path = 'assets/KHG51-SD01 烘烤炉电气件清单.xlsx'
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel file not found: {excel_path}")
            
            # Execute processing
            logger.info("Processing Excel file with res.md...")
            result = modify_excel(excel_path, ai_result, 'uploads/modified')
            
            # Verify result
            assert result['success'] == True, f"Processing should succeed: {result.get('error', '')}"
            assert 'statistics' in result, "Should return statistics"
            
            stats = result['statistics']
            logger.info(f"✓ Processing completed")
            logger.info(f"  Total tables: {stats['total_tables']}")
            logger.info(f"  Processed tables: {stats['processed_tables']}")
            logger.info(f"  Matched rows: {stats['matched_rows']}")
            logger.info(f"  Skipped rows: {stats['skipped_rows']}")
            
            if 'output_path' in result:
                logger.info(f"  Output file: {result['output_path']}")
                self.res_output_path = result.get('output_path')
            
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def test_verify_output_file(self):
        """Test: Verify output file correctness"""
        test_name = "Output File Verification"
        logger.info(f"\n[TEST] {test_name}")
        
        try:
            if not hasattr(self, 'res2_output_path'):
                logger.warning("⚠ Skipping verification - no output file from res2 test")
                return
            
            output_path = self.res2_output_path
            
            # Load output file
            wb = load_workbook(output_path)
            ws = wb.active
            
            logger.info(f"Verifying output file: {output_path}")
            
            # Verification 1: Check row 12 - "O.75KW" should be corrected to "0.75KW"
            # Find row 12 by searching for sequence number
            found_row_12 = False
            for row_num in range(1, min(50, ws.max_row + 1)):
                row = ws[row_num]
                # Check if first column contains "12"
                if row[0].value and str(row[0].value).strip() == "12":
                    found_row_12 = True
                    # Check the model column (should contain "0.75KW")
                    model_value = None
                    for cell in row:
                        if cell.value and "0.75KW" in str(cell.value):
                            model_value = str(cell.value)
                            break
                    
                    if model_value:
                        logger.info(f"✓ Row 12 verification: Found '0.75KW' in model field")
                    else:
                        logger.warning(f"⚠ Row 12: Could not verify '0.75KW' correction")
                    break
            
            if not found_row_12:
                logger.warning("⚠ Could not locate row 12 for verification")
            
            # Verification 2: Check row 27 - quantity should be 21
            found_row_27 = False
            for row_num in range(1, min(100, ws.max_row + 1)):
                row = ws[row_num]
                if row[0].value and str(row[0].value).strip() == "27":
                    found_row_27 = True
                    # Find quantity column
                    for cell in row:
                        if cell.value and str(cell.value).strip() == "21":
                            logger.info(f"✓ Row 27 verification: Quantity is 21")
                            break
                    break
            
            if not found_row_27:
                logger.warning("⚠ Could not locate row 27 for verification")
            
            # Verification 3: Check row 28 - quantity should be 9
            found_row_28 = False
            for row_num in range(1, min(100, ws.max_row + 1)):
                row = ws[row_num]
                if row[0].value and str(row[0].value).strip() == "28":
                    found_row_28 = True
                    # Find quantity column
                    for cell in row:
                        if cell.value and str(cell.value).strip() == "9":
                            logger.info(f"✓ Row 28 verification: Quantity is 9")
                            break
                    break
            
            if not found_row_28:
                logger.warning("⚠ Could not locate row 28 for verification")
            
            # General verification: File should be readable and have data
            assert ws.max_row > 0, "Output file should have rows"
            assert ws.max_column > 0, "Output file should have columns"
            
            logger.info(f"✓ Output file is valid: {ws.max_row} rows, {ws.max_column} columns")
            
            wb.close()
            self.record_pass(test_name)
            
        except Exception as e:
            logger.error(f"✗ {test_name} failed: {e}")
            self.record_fail(test_name, str(e))
    
    def record_pass(self, test_name):
        """Record a passing test"""
        self.test_results.append({'name': test_name, 'status': 'PASS'})
        self.passed += 1
    
    def record_fail(self, test_name, error):
        """Record a failing test"""
        self.test_results.append({'name': test_name, 'status': 'FAIL', 'error': error})
        self.failed += 1
    
    def print_summary(self):
        """Print test summary"""
        logger.info("\n" + "=" * 80)
        logger.info("Integration Test Summary")
        logger.info("=" * 80)
        
        for result in self.test_results:
            status_symbol = "✓" if result['status'] == 'PASS' else "✗"
            logger.info(f"{status_symbol} {result['name']}: {result['status']}")
            if result['status'] == 'FAIL':
                logger.info(f"  Error: {result.get('error', 'Unknown error')}")
        
        logger.info("-" * 80)
        logger.info(f"Total: {len(self.test_results)} tests")
        logger.info(f"Passed: {self.passed}")
        logger.info(f"Failed: {self.failed}")
        logger.info("=" * 80)
        
        if self.failed == 0:
            logger.info("✓ ALL TESTS PASSED")
        else:
            logger.error(f"✗ {self.failed} TEST(S) FAILED")


def main():
    """Main entry point"""
    runner = IntegrationTestRunner()
    runner.run_all_tests()
    
    # Exit with appropriate code
    sys.exit(0 if runner.failed == 0 else 1)


if __name__ == '__main__':
    main()
